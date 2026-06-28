"""発注管理表.xlsx を BigQuery から毎日生成して GCS へアップロード（Option A・Linux側）。

背景: 従来は Xサーバ(Windows・ZOZO IPロック)の run_daily.ps1 が生成していたが、
2026-06-19 以降そのジョブが停止し xlsx が古いまま（顧客の「常に直近7日/30日反映」未達）。
本スクリプトは ZOZO取得に依存しない BigQuery だけで生成するため、私の環境(Linux)で毎日確実に動く。

データ源:
  ・母集合 = 最新在庫スナップショット ∪ 直近30日受注 ∪ MMS発注実績(発注書一覧)
            （倉庫在庫だけだと売り切れ=在庫0の商品が漏れるため、過去に発注した商品も
              MMSから必ず拾う＝顧客要件。予約管理表は季節入替で不安定なため不採用）
  ・7日/30日販売 = sales_daily の最新受注日基準（常に最新）
  ・在庫/入荷残/予約 = 取得できている最新スナップショット（best-effort）
  ・原価 = PF手数料表(下代・品番単位)優先 → MMS評価額(SKU単位)
  ・緊急度 = 顧客定義2026: 欠品(フリー在庫≤0) / 不足(在庫日数<90) / 適正(90〜270) / 過剰(≥271)

出力: 3シート（経営者サマリ / 発注管理表 / 緊急度別 集計）→ GCS exports バケットへ。
"""
import io
import math
import sys
from datetime import datetime, timezone, timedelta

from google.cloud import bigquery, storage
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

PROJECT = "mono-back-office-system"
A = f"{PROJECT}.analytics_layer"
BUCKET = "mono-back-office-system-exports"
BLOB_LATEST = "order_management/latest/発注管理表.xlsx"
COVERAGE_WEEKS = 8           # 推奨発注 = MAX(0, ceil(8週×7×30日速度 − フリー在庫))
WARN_DAYS, OVER_DAYS = 90, 271   # 不足<90 / 過剰≥271（顧客定義2026・通年基準）

bq = bigquery.Client(project=PROJECT)

SQL = f"""
WITH
asof   AS (SELECT MAX(sale_date) d FROM `{A}.sales_daily` WHERE source_file='orders'),
inv_d  AS (SELECT MAX(snapshot_date) d FROM `{A}.inventory_snapshot`),
sa_d   AS (SELECT MAX(snapshot_date) d FROM `{A}.stock_analysis`),
inv AS (
  SELECT UPPER(TRIM(product_code)) pc, UPPER(TRIM(sku_code)) sk,
         ANY_VALUE(product_code) pc_disp, ANY_VALUE(product_name) nm, ANY_VALUE(color_name) color,
         ANY_VALUE(size) sz, ANY_VALUE(shop_name) shop,
         SUM(stock_quantity) stock, SUM(incoming_quantity) incoming, MAX(arrival_date) last_arr
  FROM `{A}.inventory_snapshot` WHERE snapshot_date=(SELECT d FROM inv_d) GROUP BY pc, sk),
fav AS (
  SELECT UPPER(TRIM(product_code)) pc, UPPER(TRIM(sku_code)) sk, SUM(favorites) fav
  FROM `{A}.stock_analysis` WHERE snapshot_date=(SELECT d FROM sa_d) GROUP BY pc, sk),
resv AS (
  SELECT UPPER(TRIM(product_code)) pc, UPPER(TRIM(sku_code)) sk, SUM(quantity) q
  FROM `{A}.reservations` WHERE reservation_date=(SELECT MAX(reservation_date) FROM `{A}.reservations`) GROUP BY pc, sk),
s7 AS (
  SELECT UPPER(TRIM(product_code)) pc, UPPER(TRIM(sku_code)) sk, SUM(sales_quantity) q
  FROM `{A}.sales_daily`, asof WHERE source_file='orders'
    AND sale_date BETWEEN DATE_SUB(asof.d, INTERVAL 6 DAY) AND asof.d GROUP BY pc, sk),
s30 AS (
  SELECT UPPER(TRIM(product_code)) pc, UPPER(TRIM(sku_code)) sk, SUM(sales_quantity) q, SUM(sales_amount) rev
  FROM `{A}.sales_daily`, asof WHERE source_file='orders'
    AND sale_date BETWEEN DATE_SUB(asof.d, INTERVAL 29 DAY) AND asof.d GROUP BY pc, sk),
rel AS (
  SELECT UPPER(TRIM(product_code)) pc, UPPER(TRIM(sku_code)) sk, MIN(sale_date) d
  FROM `{A}.sales_daily` WHERE source_file='orders' GROUP BY pc, sk),
pm AS (
  SELECT UPPER(TRIM(product_code)) pc, UPPER(TRIM(sku_code)) sk, ANY_VALUE(product_code) pc_disp,
         ANY_VALUE(product_name) nm, ANY_VALUE(color_name) color, ANY_VALUE(size) sz, ANY_VALUE(shop_name) shop
  FROM `{A}.product_master` GROUP BY pc, sk),
pf AS (
  SELECT UPPER(TRIM(product_code)) pc, ANY_VALUE(cost_price) v FROM `{A}.pf_fee_master`
  WHERE snapshot_date=(SELECT MAX(snapshot_date) FROM `{A}.pf_fee_master`) AND cost_price>0 GROUP BY pc),
mms AS (
  SELECT pc, sk, vp FROM (
    SELECT UPPER(TRIM(product_code)) pc, UPPER(TRIM(sku_code)) sk, valuation_price vp,
           ROW_NUMBER() OVER (PARTITION BY UPPER(TRIM(product_code)), UPPER(TRIM(sku_code)) ORDER BY source_date DESC) rn
    FROM `{A}.cost_master`) WHERE rn=1),
-- MMS発注実績のある品番の「商品マスタ全SKU」。過去に発注した商品を母集合に足すための種。
--   MMSのsku_codeを直接使わない理由: 発注書には sku_code 空欄の行があり(=品番単位の発注)、
--   そのまま使うと色・名称が空の幽霊行になる。商品マスタ経由なら色名が必ず付き、空欄も排除できる。
mord AS (
  SELECT DISTINCT UPPER(TRIM(pm.product_code)) pc, UPPER(TRIM(pm.sku_code)) sk
  FROM `{A}.product_master` pm
  WHERE UPPER(TRIM(pm.product_code)) IN (
    SELECT UPPER(TRIM(product_code)) FROM `{A}.mms_orders` WHERE TRIM(IFNULL(product_code,''))!='')),
universe AS (
  -- 顧客要件: 倉庫在庫を起点にすると売り切れ(在庫0)の商品が母集合から漏れる。
  --   そこで「MMSで発注実績のある品番」の商品マスタ全SKU(mord)を足し、過去に発注した
  --   商品は在庫0でも必ず・色名つきで拾う。
  --   （予約管理表は yySS/yyAW で毎シーズン入替・アーカイブされ不安定なため不採用＝顧客見解どおり）
  --   商品マスタ全件(14万)は廃番含み7倍膨張のため不採用。MMS発注のある1,089品番分だけに限定。
  SELECT pc, sk FROM inv
  UNION DISTINCT
  SELECT pc, sk FROM s30
  UNION DISTINCT
  SELECT pc, sk FROM mord)
SELECT
  COALESCE(inv.pc_disp, pm.pc_disp, u.pc) AS product_code,
  COALESCE(inv.nm, pm.nm) AS product_name,
  COALESCE(inv.color, pm.color) AS color_name,
  COALESCE(inv.sz, pm.sz) AS size,
  u.sk AS sku_code,
  COALESCE(inv.shop, pm.shop) AS shop_name,
  COALESCE(inv.stock, 0) AS stock,
  COALESCE(inv.incoming, 0) AS incoming,
  COALESCE(resv.q, 0) AS reserved,
  COALESCE(s7.q, 0) AS s7,
  COALESCE(s30.q, 0) AS s30q,
  COALESCE(s30.rev, 0) AS rev30,
  COALESCE(fav.fav, 0) AS favorites,
  CAST(inv.last_arr AS STRING) AS last_arrival,
  CAST(rel.d AS STRING) AS release_date,
  COALESCE(pf.v, mms.vp, 0) AS cost
FROM universe u
LEFT JOIN inv  ON inv.pc = u.pc  AND inv.sk = u.sk
LEFT JOIN pm   ON pm.pc = u.pc   AND pm.sk = u.sk
LEFT JOIN fav  ON fav.pc = u.pc  AND fav.sk = u.sk
LEFT JOIN resv ON resv.pc = u.pc AND resv.sk = u.sk
LEFT JOIN s7   ON s7.pc = u.pc   AND s7.sk = u.sk
LEFT JOIN s30  ON s30.pc = u.pc  AND s30.sk = u.sk
LEFT JOIN rel  ON rel.pc = u.pc  AND rel.sk = u.sk
LEFT JOIN pf   ON pf.pc = u.pc
LEFT JOIN mms  ON mms.pc = u.pc  AND mms.sk = u.sk
"""

HEADERS = ['品番', '商品名', 'カラー', 'サイズ', 'CS品番', 'ショップ', '倉庫在庫', '入荷残', '予約未処理',
           'フリー在庫', '7日販売数', '30日販売数', 'お気に入り登録数', '7日販売速度', '30日販売速度',
           '在庫日数', '最終入荷日', '推奨発注数', '確定発注数', '緊急度', '欠品まで日数', 'SKU原価',
           '30日売上', '30日原価', '粗利率(%)', '確定リリース日']


def urgency(free, stock_days):
    if free <= 0:
        return '欠品'
    if stock_days is not None and stock_days < WARN_DAYS:
        return '不足'
    if stock_days is not None and stock_days >= OVER_DAYS:
        return '過剰'
    return '適正'


def build_rows():
    rows = []
    for r in bq.query(SQL).result():
        free = (r['stock'] or 0) + (r['incoming'] or 0) - (r['reserved'] or 0)
        v7 = round((r['s7'] or 0) / 7, 2)
        v30 = round((r['s30q'] or 0) / 30, 2)
        stock_days = round(free / v7, 1) if v7 > 0 else None
        rec = max(0, math.ceil(COVERAGE_WEEKS * 7 * v30 - free))
        cost = float(r['cost'] or 0)
        rev30 = float(r['rev30'] or 0)
        cost30 = (r['s30q'] or 0) * cost
        margin = round((rev30 - cost30) / rev30 * 100, 1) if rev30 > 0 else None
        rows.append({
            'product_code': r['product_code'], 'product_name': r['product_name'], 'color': r['color_name'],
            'size': r['size'], 'sku': r['sku_code'], 'shop': r['shop_name'],
            'stock': r['stock'] or 0, 'incoming': r['incoming'] or 0, 'reserved': r['reserved'] or 0,
            'free': free, 's7': r['s7'] or 0, 's30': r['s30q'] or 0, 'fav': r['favorites'] or 0,
            'v7': v7, 'v30': v30, 'stock_days': stock_days, 'last_arr': r['last_arrival'],
            'rec': rec, 'urg': urgency(free, stock_days), 'stockout_days': stock_days,
            'cost': cost, 'rev30': rev30, 'cost30': cost30, 'margin': margin, 'release': r['release_date'],
        })
    return rows


def write_sheet_detail(ws, rows):
    bold = Font(bold=True, color='FFFFFF')
    fill = PatternFill('solid', fgColor='374151')
    urg_fill = {'欠品': 'FCA5A5', '不足': 'FDE68A', '過剰': 'BFDBFE', '適正': None}
    ws.append(HEADERS)
    for c in ws[1]:
        c.font = bold; c.fill = fill; c.alignment = Alignment(horizontal='center')
    for r in rows:
        ws.append([
            r['product_code'], r['product_name'], r['color'], r['size'], r['sku'], r['shop'],
            r['stock'], r['incoming'], r['reserved'], r['free'], r['s7'], r['s30'], r['fav'],
            r['v7'], r['v30'], r['stock_days'], r['last_arr'], r['rec'], '', r['urg'], r['stockout_days'],
            round(r['cost']), round(r['rev30']), round(r['cost30']),
            r['margin'], r['release'],
        ])
        f = urg_fill.get(r['urg'])
        if f:
            ws.cell(ws.max_row, 20).fill = PatternFill('solid', fgColor=f)
    ws.freeze_panes = 'A2'
    widths = [10, 22, 10, 6, 9, 12, 8, 7, 9, 9, 9, 10, 11, 11, 11, 9, 11, 9, 9, 7, 10, 8, 11, 11, 8, 11]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w


def write_sheet_urgent(ws, rows):
    top = sorted([r for r in rows if r['rec'] > 0], key=lambda x: -x['rec'])[:200]
    ws.append(['緊急度別 SKU 上位 (推奨発注数 降順・上位200)'])
    ws.append([])
    hdr = ['緊急度', '品番', 'ショップ', '商品名', 'カラー', 'サイズ', 'フリー在庫', '推奨発注数', '推奨発注金額', '30日販売速度']
    ws.append(hdr)
    bold = Font(bold=True)
    for c in ws[3]:
        c.font = bold
    for r in top:
        ws.append([r['urg'], r['product_code'], r['shop'], r['product_name'], r['color'], r['size'],
                   r['free'], r['rec'], round(r['rec'] * r['cost']), r['v30']])
    for i, w in enumerate([8, 10, 12, 22, 10, 6, 10, 10, 14, 12], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w


def write_sheet_summary(ws, rows, asof, inv_date):
    n = len(rows)
    by_urg = {u: sum(1 for r in rows if r['urg'] == u) for u in ('欠品', '不足', '適正', '過剰')}
    total_rec = sum(r['rec'] for r in rows)
    total_rec_amt = sum(r['rec'] * r['cost'] for r in rows)
    stock_amt = sum(r['stock'] * r['cost'] for r in rows)
    margins = [r['margin'] for r in rows if r['margin'] is not None]
    avg_margin = round(sum(margins) / len(margins), 1) if margins else 0
    jst = datetime.now(timezone(timedelta(hours=9)))
    rows_out = [
        [f'📊 発注管理表 — 経営者サマリ（受注最新 {asof} ・ 在庫 {inv_date}）'],
        [f'生成日時: {jst.strftime("%Y-%m-%d %H:%M JST")}（BigQueryから自動生成・常に最新の7日/30日実績）'],
        [],
        ['全SKU数', f'{n:,}'],
        ['緊急度 欠品', f'{by_urg["欠品"]:,} SKU', 'フリー在庫 ≤ 0'],
        ['緊急度 不足', f'{by_urg["不足"]:,} SKU', '在庫日数 < 90日（通年基準）'],
        ['緊急度 適正', f'{by_urg["適正"]:,} SKU', '在庫日数 90〜270日'],
        ['緊急度 過剰', f'{by_urg["過剰"]:,} SKU', '在庫日数 ≥ 271日'],
        ['推奨発注 合計数量', f'{total_rec:,} 点'],
        ['推奨発注 合計金額(原価)', f'¥{round(total_rec_amt):,}', '推奨数 × SKU原価(PF優先)'],
        ['在庫金額(原価ベース)', f'¥{round(stock_amt):,}', '在庫数 × SKU原価'],
        ['平均粗利率', f'{avg_margin}%', '全SKUの平均'],
        [],
        ['対象品番の条件', '最新在庫に載るSKU ∪ 直近30日に受注のあるSKU（売上0/在庫0も含む）'],
        ['注意', '在庫データはZOZO取得状況に依存（取得停止時は在庫が減ります）。7日/30日販売は常に最新。'],
        [],
        ['ショップ別ブレイクダウン'],
        ['ショップ', 'SKU数', '欠品', '不足', '推奨発注数', '推奨発注金額', '在庫金額', '平均粗利率'],
    ]
    for ro in rows_out:
        ws.append(ro)
    shops = {}
    for r in rows:
        s = r['shop'] or '(unknown)'
        d = shops.setdefault(s, {'n': 0, 'c': 0, 'w': 0, 'rec': 0, 'amt': 0, 'stk': 0, 'm': []})
        d['n'] += 1; d['rec'] += r['rec']; d['amt'] += r['rec'] * r['cost']; d['stk'] += r['stock'] * r['cost']
        if r['urg'] == '欠品': d['c'] += 1
        if r['urg'] == '不足': d['w'] += 1
        if r['margin'] is not None: d['m'].append(r['margin'])
    for s, d in sorted(shops.items(), key=lambda x: -x[1]['n']):
        am = round(sum(d['m']) / len(d['m']), 1) if d['m'] else 0
        ws.append([s, d['n'], d['c'], d['w'], d['rec'], round(d['amt']), round(d['stk']), f'{am}%'])
    ws['A1'].font = Font(bold=True, size=13)
    ws.column_dimensions['A'].width = 22
    for col in 'BCDEFGH':
        ws.column_dimensions[col].width = 16


def main():
    rows = build_rows()
    asof = list(bq.query(f"SELECT CAST(MAX(sale_date) AS STRING) d FROM `{A}.sales_daily` WHERE source_file='orders'"))[0]['d']
    inv_date = list(bq.query(f"SELECT CAST(MAX(snapshot_date) AS STRING) d FROM `{A}.inventory_snapshot`"))[0]['d']
    print(f"rows={len(rows)} / 受注最新={asof} / 在庫={inv_date}")

    wb = openpyxl.Workbook()
    write_sheet_summary(wb.active, rows, asof, inv_date)
    wb.active.title = '経営者サマリ'
    write_sheet_detail(wb.create_sheet('発注管理表'), rows)
    write_sheet_urgent(wb.create_sheet('緊急度別 集計'), rows)

    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()

    sc = storage.Client(project=PROJECT)
    bkt = sc.bucket(BUCKET)
    bkt.blob(BLOB_LATEST).upload_from_string(
        data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    dated = f"order_management/{datetime.now(timezone(timedelta(hours=9))).strftime('%Y%m%d')}/発注管理表.xlsx"
    bkt.blob(dated).upload_from_string(
        data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    print(f"uploaded gs://{BUCKET}/{BLOB_LATEST}  ({len(data):,}B)  + {dated}")


if __name__ == '__main__':
    sys.exit(main())
