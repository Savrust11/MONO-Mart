import sys, io, csv
from pathlib import Path
sys.path.insert(0, r"C:\Users\Administrator\Downloads\system\pipeline")
KEY = Path(r"C:\Users\Administrator\Downloads\system\pipeline\sheets-sa-key.json")
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from openpyxl import load_workbook

SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]
creds = Credentials.from_service_account_file(str(KEY), scopes=SCOPES)
drive = build("drive", "v3", credentials=creds, cache_discovery=False)

def find(fid, name):
    res = drive.files().list(q=f"'{fid}' in parents and trashed=false and name='{name}'",
        fields="files(id,name)", supportsAllDrives=True, includeItemsFromAllDrives=True).execute()
    return res.get("files", [])
def dl(fid):
    buf = io.BytesIO(); req = drive.files().get_media(fileId=fid, supportsAllDrives=True)
    d = MediaIoBaseDownload(buf, req); done=False
    while not done: _, done = d.next_chunk()
    return buf.getvalue()
def col_letter(i):
    s=""; n=i
    while True:
        s=chr(ord("A")+n%26)+s; n=n//26-1
        if n<0: break
    return s
def header_xlsx(data):
    wb=load_workbook(io.BytesIO(data), read_only=True); ws=wb.active
    return [c for c in next(ws.iter_rows(min_row=1,max_row=1,values_only=True))]
def header_csv(data):
    for enc in ("utf-8-sig","cp932","shift_jis"):
        try: text=data.decode(enc); break
        except UnicodeDecodeError: continue
    return next(csv.reader(io.StringIO(text)))

def report(hdr, label):
    print(f"\n=== {label} — {len(hdr)} columns ===")
    for i,name in enumerate(hdr):
        nm=str(name).strip() if name else ""
        mark=""
        if any(k in nm for k in ("発送日","出荷日","注文日","売上日","日")): mark=" <== DATE?"
        if nm in ("CS品番","ブランド品番","注文数","出荷数","数量","合計金額（税抜）"): mark+=" <== KEY"
        print(f"  {col_letter(i):>2}({i+1:2d}) {nm}{mark}")

mf=find("1eRZoby0Bo8J4c6anWoe0PSGpDaGjePaR","202407")
of=find(mf[0]["id"],"2024_07_01.xlsx")
report(header_xlsx(dl(of[0]["id"])), "ORDERS 2024_07_01.xlsx")

sx=find("1kSuIgbkOtpfOu474A-rFRWKcLN0wKjJK","20240701(出荷).xlsx")
report(header_xlsx(dl(sx[0]["id"])), "SHIPPED 20240701(出荷).xlsx")

scv=find("1kSuIgbkOtpfOu474A-rFRWKcLN0wKjJK","20240801(出荷).csv")
report(header_csv(dl(scv[0]["id"])), "SHIPPED 20240801(出荷).csv")
