from openpyxl import load_workbook
wb = load_workbook(r"C:\Users\Administrator\Downloads\system\exports\発注管理表_sc1032_2026-05-01_2026-05-31.xlsx")
ws = wb.active
for row in ws.iter_rows(values_only=True):
    label, val, note = (list(row)+["","",""])[:3]
    label = "" if label is None else str(label)
    val = "" if val is None else str(val)
    note = "" if note is None else str(note)
    print(f"{label:<24} | {val:<14} | {note}")
