"""発注管理表 Excel生成 (Phase 2 / 2026-06-03 v1)

mart_layer.order_analysis の最新日付から、クライアント仕様
(docs/order_management_sheet_spec.md §4) の 23 列を使い、3 シートの xlsx を
生成して GCS にアップロードする:

  Sheet1 「経営者サマリ」  — 集計日 / SKU数 / 緊急度別件数 / 推奨発注金額・在庫金額・
                              平均粗利率 (全体 + ショップ別ブレイクダウン)
  Sheet2 「発注管理表」    — 24 列 SKU 詳細 (緊急度別の色分け)
  Sheet3 「緊急度別 集計」  — CRITICAL / WARNING / OK / OVERSTOCK の合計サマリ

出力先 (公開 URL — クライアントは bookmark 1 つで毎日の最新版にアクセス可):
  gs://mono-back-office-system-exports/order_management/latest/発注管理表.xlsx
  gs://mono-back-office-system-exports/order_management/{date}/発注管理表_yyyymmdd.xlsx

ENV:
  TARGET_DATE      (default: mart の MAX(analysis_date))
  GCS_EXPORTS_BUCKET (default: mono-back-office-system-exports)
"""
from __future__ import annotations

import os, sys, io, logging
from datetime import date as date_cls, datetime, timezone, timedelta
from pathlib import Path
from typing import Any

from google.cloud import bigquery, storage

# openpyxl is bundled with the existing python env (used by xlsx imports
# elsewhere). Import here so missing-package errors surface early.
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("order_excel")

JST = timezone(timedelta(hours=9))
PROJECT = "mono-back-office-system"
EXPORTS_BUCKET = os.getenv("GCS_EXPORTS_BUCKET", "mono-back-office-system-exports")

# ── スタイル ──────────────────────────────────────────────────────────────
THIN = Side(border_style="thin", color="D0D5DD")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FONT = Font(name="Meiryo", size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="475467")
BODY_FONT = Font(name="Meiryo", size=10)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")

# 緊急度別の行背景色
URGENCY_FILLS = {
    "CRITICAL":  PatternFill("solid", fgColor="FEE4E2"),   # 赤系
    "WARNING":   PatternFill("solid", fgColor="FEF3C7"),   # 黄系
    "OK":        PatternFill("solid", fgColor="D1FADF"),   # 緑系
    "OVERSTOCK": PatternFill("solid", fgColor="DBEAFE"),   # 青系
}

# 仕様書 §4 の 24 列 (確定リリース日は join できる場合のみ)
SHEET2_COLUMNS: list[tuple[str, str]] = [
    ("品番",           "product_code"),
    ("商品名",         "product_name"),
    ("カラー",         "color_name"),
    ("サイズ",         "size"),
    ("CS品番",         "sku_code"),
    ("ショップ",       "shop_name"),
    ("倉庫在庫",       "inventory"),
    ("入荷残",         "incoming_stock"),
    ("予約未処理",     "reservations_pending"),
    ("フリー在庫",     "free_inventory"),
    ("7日販売数",      "sales_7d"),
    ("30日販売数",     "sales_30d"),
    ("お気に入り登録数", "favorites_total"),
    ("7日販売速度",    "daily_velocity_7d"),
    ("30日販売速度",   "daily_velocity_30d"),
    ("在庫日数",       "stock_days_7d"),
    ("最終入荷日",     "arrival_date"),
    ("推奨発注数",     "recommended_order_qty"),
    ("確定発注数",     "confirmed_order_qty"),
    ("緊急度",         "order_urgency"),
    ("欠品まで日数",   "days_to_stockout"),
    ("SKU原価",        "cost_price"),
    ("30日売上",       "period_revenue"),
    ("30日原価",       "period_total_cost"),
    ("粗利率(%)",      "gross_margin_pct"),
    ("確定リリース日", "confirmed_release_date"),
]


def _get_target_date(bq: bigquery.Client) -> str:
    env = os.getenv("TARGET_DATE")
    if env:
        return env
    r = list(bq.query(
        "SELECT MAX(analysis_date) AS d FROM `mart_layer.order_analysis`"
    ).result())[0]
    if r.d is None:
        raise RuntimeError("mart_layer.order_analysis is empty")
    return r.d.isoformat()


def _fetch_rows(bq: bigquery.Client, target_date: str) -> list[dict]:
    sql = """
    SELECT
      o.product_code, o.product_name, o.color_name, o.size, o.sku_code,
      pm.shop_name,
      o.inventory, o.incoming_stock, o.reservations_pending, o.free_inventory,
      o.sales_7d, o.sales_30d, o.favorites_total, o.daily_velocity_7d, o.daily_velocity_30d,
      o.stock_days_7d, o.recommended_order_qty, o.order_urgency,
      o.days_to_stockout, o.cost_price, o.period_revenue, o.period_total_cost,
      o.gross_margin_pct, o.arrival_date,
      st.confirmed_release_date
    FROM `mart_layer.order_analysis` o
    LEFT JOIN (
      SELECT product_code, sku_code, ANY_VALUE(shop_name) AS shop_name
      FROM `analytics_layer.product_master`
      GROUP BY product_code, sku_code
    ) pm
      ON pm.product_code = o.product_code AND pm.sku_code = o.sku_code
    LEFT JOIN (
      SELECT product_code, ANY_VALUE(confirmed_release_date) AS confirmed_release_date
      FROM `analytics_layer.sitateru_item_master`
      WHERE confirmed_release_date IS NOT NULL
      GROUP BY product_code
    ) st
      ON st.product_code = o.product_code
    WHERE o.analysis_date = @d
    ORDER BY
      CASE o.order_urgency
        WHEN 'CRITICAL' THEN 1 WHEN 'WARNING' THEN 2
        WHEN 'OK' THEN 3 WHEN 'OVERSTOCK' THEN 4 ELSE 5
      END,
      o.recommended_order_qty DESC,
      o.product_code, o.sku_code
    """
    rows = list(bq.query(sql, job_config=bigquery.QueryJobConfig(
        query_parameters=[bigquery.ScalarQueryParameter("d", "DATE", target_date)]
    )).result())
    return [dict(r.items()) for r in rows]


def _build_summary_sheet(ws, target_date: str, rows: list[dict]) -> None:
    """Sheet1 経営者サマリ — 集計日 / SKU数 / 緊急度件数 / 推奨発注金額 等"""
    ws.title = "経営者サマリ"
    ws.sheet_view.showGridLines = False

    title = ws.cell(row=1, column=1,
                    value=f"📊 発注管理表 — 経営者サマリ ({target_date})")
    title.font = Font(name="Meiryo", size=16, bold=True, color="101828")
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 26

    info = ws.cell(row=2, column=1,
                   value=f"生成日時: {datetime.now(JST).strftime('%Y-%m-%d %H:%M JST')}")
    info.font = Font(name="Meiryo", size=10, color="667085")
    ws.merge_cells("A2:F2")

    # 全体KPI
    total_skus = len(rows)
    urgency_counts = {"CRITICAL": 0, "WARNING": 0, "OK": 0, "OVERSTOCK": 0}
    rec_total_qty = 0
    rec_total_amount = 0
    inv_total_amount = 0
    period_revenue_sum = 0.0
    period_cost_sum = 0.0
    for r in rows:
        u = r.get("order_urgency") or "OK"
        urgency_counts[u] = urgency_counts.get(u, 0) + 1
        q = r.get("recommended_order_qty") or 0
        c = r.get("cost_price") or 0
        inv = r.get("inventory") or 0
        rec_total_qty += int(q)
        rec_total_amount += int(q) * float(c)
        inv_total_amount += int(inv) * float(c)
        period_revenue_sum += float(r.get("period_revenue") or 0)
        period_cost_sum += float(r.get("period_total_cost") or 0)
    avg_margin = (1.0 - period_cost_sum / period_revenue_sum) * 100 if period_revenue_sum > 0 else 0.0

    # KPI table
    kpi_start = 4
    kpi_rows = [
        ("全SKU数",               f"{total_skus:,}",                     ""),
        ("緊急度 CRITICAL",       f"{urgency_counts['CRITICAL']:,} SKU", "欠品リスク (フリー在庫 ≤ 0)"),
        ("緊急度 WARNING",        f"{urgency_counts['WARNING']:,} SKU",  "在庫日数 ≤ 14日"),
        ("緊急度 OK",             f"{urgency_counts['OK']:,} SKU",       "適正水準"),
        ("緊急度 OVERSTOCK",      f"{urgency_counts['OVERSTOCK']:,} SKU","在庫日数 > 90日"),
        ("推奨発注 合計数量",     f"{rec_total_qty:,} 点",               ""),
        ("推奨発注 合計金額(原価)", f"¥{int(rec_total_amount):,}",         "推奨数 × SKU原価"),
        ("在庫金額(原価ベース)",   f"¥{int(inv_total_amount):,}",         "在庫数 × SKU原価"),
        ("平均粗利率",            f"{avg_margin:.1f}%",                 "全SKUの平均"),
    ]
    for i, (label, val, note) in enumerate(kpi_rows):
        r = kpi_start + i
        for c, v, fnt, fill, align in (
            (1, label, Font(name="Meiryo", size=11, bold=True, color="344054"),
             None, ALIGN_LEFT),
            (2, val, Font(name="Meiryo", size=11, bold=True, color="101828"),
             None, ALIGN_RIGHT),
            (3, note, Font(name="Meiryo", size=9, color="667085"),
             None, ALIGN_LEFT),
        ):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = fnt
            cell.alignment = align
            if fill:
                cell.fill = fill

    # ショップ別ブレイクダウン
    shops: dict[str, dict] = {}
    for r in rows:
        s = r.get("shop_name") or "(unknown)"
        x = shops.setdefault(s, {"skus": 0, "crit": 0, "warn": 0,
                                 "rec_qty": 0, "rec_amt": 0.0,
                                 "inv_amt": 0.0, "rev_sum": 0.0,
                                 "cost_sum": 0.0})
        x["skus"] += 1
        u = r.get("order_urgency") or "OK"
        if u == "CRITICAL":
            x["crit"] += 1
        elif u == "WARNING":
            x["warn"] += 1
        q = int(r.get("recommended_order_qty") or 0)
        c = float(r.get("cost_price") or 0)
        inv = int(r.get("inventory") or 0)
        x["rec_qty"] += q
        x["rec_amt"] += q * c
        x["inv_amt"] += inv * c
        x["rev_sum"] += float(r.get("period_revenue") or 0)
        x["cost_sum"] += float(r.get("period_total_cost") or 0)

    breakdown_start = kpi_start + len(kpi_rows) + 2
    ws.cell(row=breakdown_start, column=1, value="ショップ別ブレイクダウン").font = (
        Font(name="Meiryo", size=13, bold=True, color="101828"))

    bd_headers = ["ショップ", "SKU数", "CRITICAL", "WARNING",
                  "推奨発注数", "推奨発注金額", "在庫金額", "平均粗利率"]
    hdr_row = breakdown_start + 2
    for i, h in enumerate(bd_headers, start=1):
        cell = ws.cell(row=hdr_row, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_CENTER
        cell.border = BOX

    for j, (shop, x) in enumerate(sorted(
            shops.items(), key=lambda kv: -kv[1]["rec_amt"])):
        row = hdr_row + 1 + j
        am = (1.0 - x["cost_sum"] / x["rev_sum"]) * 100 if x["rev_sum"] > 0 else 0.0
        values = [
            shop, x["skus"], x["crit"], x["warn"],
            x["rec_qty"], int(x["rec_amt"]), int(x["inv_amt"]),
            f"{am:.1f}%",
        ]
        for i, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=i, value=v)
            cell.font = BODY_FONT
            cell.border = BOX
            cell.alignment = ALIGN_RIGHT if i > 1 else ALIGN_LEFT
            if i in (6, 7):  # 金額列
                cell.number_format = "¥#,##0"
            elif i in (2, 3, 4, 5):
                cell.number_format = "#,##0"

    # Column widths
    for col, w in zip("ABCDEFGH", (24, 22, 30, 0, 0, 0, 0, 0)):
        if w:
            ws.column_dimensions[col].width = w
    for col in "BCDEFGH":
        if ws.column_dimensions[col].width is None or ws.column_dimensions[col].width < 14:
            ws.column_dimensions[col].width = 14


def _build_detail_sheet(ws, rows: list[dict]) -> None:
    """Sheet2 発注管理表 — 23 列 SKU 詳細 + 緊急度色分け"""
    ws.title = "発注管理表"
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    # Header row
    for i, (label, _key) in enumerate(SHEET2_COLUMNS, start=1):
        cell = ws.cell(row=1, column=i, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_CENTER
        cell.border = BOX

    # 列順 (2026-06-19 最終入荷日=17, 確定発注数=19 を追加):
    # 7-13 在庫/販売/お気に入り, 14-16 速度/在庫日数, 17 最終入荷日,
    # 18 推奨発注数, 19 確定発注数(手入力・空), 20 緊急度, 21 欠品まで日数,
    # 22-24 原価/売上/30日原価, 25 粗利率, 26 確定リリース日
    int_cols = {7, 8, 9, 10, 11, 12, 13, 18, 21}
    money_cols = {22, 23, 24}
    float_cols = {14, 15, 16}
    pct_cols = {25}

    for r_idx, r in enumerate(rows, start=2):
        urgency = (r.get("order_urgency") or "OK").upper()
        fill = URGENCY_FILLS.get(urgency)
        for c_idx, (_label, key) in enumerate(SHEET2_COLUMNS, start=1):
            val = r.get(key)
            cell = ws.cell(row=r_idx, column=c_idx)
            # Convert types for display
            if isinstance(val, (datetime, date_cls)):
                val = val.isoformat() if val else ""
            if c_idx in pct_cols and val is not None:
                cell.value = float(val)
                cell.number_format = "0.0\"%\""
            elif c_idx in money_cols and val is not None:
                cell.value = float(val)
                cell.number_format = "¥#,##0"
            elif c_idx in float_cols and val is not None:
                cell.value = float(val)
                cell.number_format = "0.00"
            elif c_idx in int_cols and val is not None:
                cell.value = int(val)
                cell.number_format = "#,##0"
            else:
                cell.value = val if val is not None else ""
            cell.font = BODY_FONT
            cell.border = BOX
            cell.alignment = (ALIGN_RIGHT
                              if (c_idx in int_cols | money_cols | float_cols | pct_cols)
                              else ALIGN_LEFT)
            if fill:
                cell.fill = fill

    # Column widths
    widths = [16, 28, 12, 8, 16, 14, 8, 8, 10, 11, 9, 10, 14, 12, 13, 10,
              12, 12, 11, 11, 12, 11, 14, 14, 11, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Auto-filter on the header row
    ws.auto_filter.ref = (
        f"A1:{get_column_letter(len(SHEET2_COLUMNS))}{max(2, len(rows)+1)}")


def _build_urgency_sheet(ws, rows: list[dict]) -> None:
    ws.title = "緊急度別 集計"
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value="緊急度別 SKU 上位 (推奨発注数 降順)").font = (
        Font(name="Meiryo", size=14, bold=True, color="101828"))

    headers = ["緊急度", "品番", "ショップ", "商品名", "カラー", "サイズ",
               "フリー在庫", "推奨発注数", "推奨発注金額", "30日販売速度"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_CENTER
        cell.border = BOX

    # Order: CRITICAL → WARNING → OK → OVERSTOCK, top 50 per
    out_rows: list[dict] = []
    for u in ("CRITICAL", "WARNING", "OK", "OVERSTOCK"):
        bucket = [r for r in rows if (r.get("order_urgency") or "OK") == u]
        bucket.sort(key=lambda x: -(x.get("recommended_order_qty") or 0))
        out_rows.extend(bucket[:50])

    for j, r in enumerate(out_rows, start=4):
        urgency = (r.get("order_urgency") or "OK").upper()
        fill = URGENCY_FILLS.get(urgency)
        q = int(r.get("recommended_order_qty") or 0)
        c = float(r.get("cost_price") or 0)
        amount = q * c
        values = [
            urgency, r.get("product_code"), r.get("shop_name"),
            r.get("product_name"), r.get("color_name"), r.get("size"),
            r.get("free_inventory"), q, amount, r.get("daily_velocity_30d"),
        ]
        for i, v in enumerate(values, start=1):
            cell = ws.cell(row=j, column=i, value=v)
            cell.font = BODY_FONT
            cell.border = BOX
            cell.alignment = ALIGN_RIGHT if i > 6 else ALIGN_LEFT
            if i == 7:
                cell.number_format = "#,##0"
            elif i == 8:
                cell.number_format = "#,##0"
            elif i == 9:
                cell.number_format = "¥#,##0"
            elif i == 10 and isinstance(v, (int, float)):
                cell.number_format = "0.00"
            if fill:
                cell.fill = fill

    for i, w in enumerate(
            [12, 14, 14, 28, 12, 8, 11, 12, 14, 13], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main() -> int:
    bq = bigquery.Client(project=PROJECT)
    target_date = _get_target_date(bq)
    logger.info("Excel build start (target=%s)", target_date)

    rows = _fetch_rows(bq, target_date)
    logger.info("Fetched %d rows from mart_layer.order_analysis", len(rows))
    if not rows:
        logger.warning("no data — exit 0")
        return 0

    wb = Workbook()
    _build_summary_sheet(wb.active, target_date, rows)
    _build_detail_sheet(wb.create_sheet(), rows)
    _build_urgency_sheet(wb.create_sheet(), rows)

    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    yyyymmdd = target_date.replace("-", "")
    filename = f"発注管理表_{yyyymmdd}.xlsx"

    sc = storage.Client(project=PROJECT)
    bkt = sc.bucket(EXPORTS_BUCKET)

    # Dated copy + "latest" stable URL
    dated_key = f"order_management/{target_date}/{filename}"
    latest_key = "order_management/latest/発注管理表.xlsx"
    ct = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    for key in (dated_key, latest_key):
        blob = bkt.blob(key)
        blob.upload_from_string(data, content_type=ct)
        try:
            blob.make_public()
        except Exception:
            pass
    logger.info("✓ Uploaded (%s bytes):\n  %s\n  %s",
                f"{len(data):,}",
                bkt.blob(dated_key).public_url,
                bkt.blob(latest_key).public_url)
    return 0


if __name__ == "__main__":
    sys.exit(main())
