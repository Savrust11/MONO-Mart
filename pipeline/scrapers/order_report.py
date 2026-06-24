# -*- coding: utf-8 -*-
"""
発注管理表（項目詳細）レポート生成 — 品番 × 指定期間 の「実数」を出力。
顧客最優先機能。確定仕様（社内で決めること）反映済み。出力先＝Googleスプレッドシート(確定)。

────────────────────────────────────────────────────────────────────────
【保守のしかた / 仕様変更時はここを見る】
  ・項目の 追加/削除/並べ替え/改名      → REPORT_LAYOUT を編集（1行）
  ・計算式の変更                          → compute_values() 内の該当キー（仕様行番号を注記）
  ・在庫日数の分母・集計窓・原価優先 等   → CONFIG を編集
  ・出力先（スプシ/xlsx）                 → main の render 呼び出し
データ未取得の項目は値 None → 出力時「（データ未取得）」と注記。勝手な計算はしない。
────────────────────────────────────────────────────────────────────────

使い方:
  # xlsx（即時レビュー用）
  python order_report.py <品番> <開始日> <終了日>
  # Googleスプレッドシート（確定仕様。事前にSAへ共有したシートIDをenvで指定）
  GSHEET_ID=xxxx python order_report.py <品番> <開始日> <終了日>
  SA: sheets-fetcher@mono-back-office-system.iam.gserviceaccount.com を編集者に追加すること
"""
import os, sys, datetime, statistics
from google.cloud import bigquery

# ============================ CONFIG（要件変更はまずここ）============================
PROJECT          = "mono-back-office-system"
SA_KEY           = r"C:\Users\Administrator\Downloads\system\pipeline\sheets-sa-key.json"
STOCK_DAYS_BASIS = "median"     # 在庫日数の分母: "median"(30日中央値) / "average"(30日平均)
WIN_SHORT        = 7            # ▼直近◯日（短期）
WIN_LONG         = 30           # ▼直近◯日（長期）
COST_PRIORITY    = ("pf", "mms")  # 原価: PF手数料表(品番)→MMS評価額(SKU)  ← 仕様R14/R69
NA               = "（データ未取得）"
OUT_DIR          = r"C:\Users\Administrator\Downloads\system\exports"
# ===================================================================================

bq = bigquery.Client(project=PROJECT)

def _p(**kw):
    return bigquery.QueryJobConfig(query_parameters=[
        bigquery.ScalarQueryParameter(k, t, v) for k, (t, v) in kw.items()])
def q1(sql, **kw):
    r = list(bq.query(sql, job_config=_p(**kw)).result()); return r[0] if r else None
def qn(sql, **kw):
    return list(bq.query(sql, job_config=_p(**kw)).result())
def asdate(v):
    if v is None or v == "": return None
    return v.isoformat() if hasattr(v, "isoformat") else str(v)


# ============================ 計算層（仕様の計算式どおり）============================
def compute_values(pc, start, end):
    """品番 pc × 期間[start,end] の全実数を計算して dict で返す。キーは REPORT_LAYOUT が参照。"""
    V = {"品番": pc, "集計開始日": start, "集計終了日": end}

    # 作成日（自動）= データ最新日を超えない当日
    today = datetime.date.today().isoformat()
    asof = q1("SELECT LEAST(@t, CAST(MAX(sale_date) AS STRING)) d "
              "FROM `analytics_layer.sales_daily` WHERE source_file='orders'",
              t=("STRING", today)).d
    asof_d = datetime.date.fromisoformat(asof)
    V["作成日"] = asof
    dL = (asof_d - datetime.timedelta(days=WIN_LONG - 1)).isoformat()

    # ── 受注期間集計（SKU単位）＋原価（PF品番→MMS SKU）R14/R15/R25/R71/R73 ──
    sku = qn("""
    WITH s AS (
      SELECT sku_code, color_name, size, sale_type,
        SUM(sales_quantity) qty, SUM(sales_amount) rev,
        SUM(proper_price*sales_quantity) lst
      FROM `analytics_layer.sales_daily`
      WHERE product_code=@pc AND source_file='orders'
        AND sale_date BETWEEN @sd AND @ed AND sales_quantity>0
      GROUP BY sku_code, color_name, size, sale_type ),
    pfc AS (
      SELECT ANY_VALUE(cost_price) val FROM `analytics_layer.pf_fee_master`
      WHERE UPPER(TRIM(product_code))=UPPER(TRIM(@pc))
        AND snapshot_date=(SELECT MAX(snapshot_date) FROM `analytics_layer.pf_fee_master`)
        AND cost_price>0 ),
    mms AS (
      SELECT j_col, j_sz, ARRAY_AGG(cost_price ORDER BY source_date DESC LIMIT 1)[OFFSET(0)] cp
      FROM (SELECT UPPER(TRIM(color_name)) j_col, UPPER(TRIM(size)) j_sz, cost_price, source_date
            FROM `analytics_layer.cost_master`
            WHERE UPPER(TRIM(product_code))=UPPER(TRIM(@pc)) AND color_name IS NOT NULL
              AND size IS NOT NULL AND valid_to IS NULL)
      GROUP BY j_col, j_sz )
    SELECT s.color_name, s.sale_type, s.qty, s.rev, s.lst,
      COALESCE((SELECT val FROM pfc), mms.cp) AS cost
    FROM s LEFT JOIN mms ON mms.j_col=UPPER(TRIM(s.color_name)) AND mms.j_sz=UPPER(TRIM(s.size))
    """, pc=("STRING", pc), sd=("DATE", start), ed=("DATE", end))

    tq  = sum(r.qty for r in sku) or 0
    trv = sum(float(r.rev) for r in sku) or 0.0
    tls = sum(float(r.lst) for r in sku) or 0.0
    tcs = sum(float(r.cost) * r.qty for r in sku if r.cost is not None)
    yq  = sum(r.qty for r in sku if r.sale_type and "予約" in str(r.sale_type))
    V["原価欠落SKU"] = sum(1 for r in sku if r.cost is None)
    V["合計販売数"]   = tq
    V["平均売価"]     = round(trv / tq, 1) if tq else None                 # R71
    V["値引率"]       = round((1 - trv / tls) * 100, 1) if tls else None   # R15
    V["粗利率"]       = round((trv - tcs) / trv * 100, 1) if trv else None # R14
    V["最新加重平均原価"] = round(tcs / tq, 1) if tq and tcs else None
    V["予約販売数割合"] = round(yq / tq * 100, 1) if tq else None          # R73
    # FKU枚数構成（品番&カラー）R25
    fku = {}
    for r in sku: fku[r.color_name] = fku.get(r.color_name, 0) + r.qty
    V["_fku"] = {c: (round(n / tq * 100, 1), n) for c, n in fku.items()} if tq else {}

    # ── マスタ属性（品番）R07-R11,R20 ──
    m = q1("""SELECT ANY_VALUE(product_name) nm, ANY_VALUE(shop_name) shop,
      ANY_VALUE(parent_category) brand, ANY_VALUE(parent_item_type) pit,
      ANY_VALUE(child_item_type) cit, ANY_VALUE(proper_price) jodai
      FROM `analytics_layer.product_master`
      WHERE UPPER(TRIM(product_code))=UPPER(TRIM(@pc))""", pc=("STRING", pc))
    V["商品名"] = m.nm if m else None
    V["ショップ"] = m.shop if m else None
    V["ブランド"] = m.brand if m else None
    V["商品タイプ親"] = m.pit if m else None
    V["商品タイプ子"] = m.cit if m else None
    V["上代"] = float(m.jodai) if m and m.jodai else None

    # 販売開始日＝期間内初回受注日（確定№2）
    sd0 = q1("""SELECT MIN(sale_date) d FROM `analytics_layer.sales_daily`
      WHERE product_code=@pc AND source_file='orders'
        AND sale_date BETWEEN @sd AND @ed AND sales_quantity>0""",
             pc=("STRING", pc), sd=("DATE", start), ed=("DATE", end))
    V["販売開始日"] = asdate(sd0.d) if sd0 else None

    # 在庫分析daily（現在）: 現在庫=S列販売可能数・予約0（確定№1）/ お気に入り（確定№11）
    st = q1("""
    WITH pm AS (SELECT UPPER(TRIM(product_code)) pc, UPPER(TRIM(sku_code)) sk,
                       ANY_VALUE(sale_type) stype FROM `analytics_layer.product_master` GROUP BY pc, sk)
    SELECT SUM(CASE WHEN pm.stype LIKE '%予約%' THEN 0 ELSE sa.available_qty END) cur,
           SUM(sa.favorites) fav
    FROM `analytics_layer.stock_analysis` sa
    LEFT JOIN pm ON pm.pc=UPPER(TRIM(sa.product_code)) AND pm.sk=UPPER(TRIM(sa.sku_code))
    WHERE sa.product_code=@pc
      AND sa.snapshot_date=(SELECT MAX(snapshot_date) FROM `analytics_layer.stock_analysis`
                            WHERE snapshot_date<=@asof)""",
            pc=("STRING", pc), asof=("DATE", asof))
    cur = (st.cur if st and st.cur is not None else 0)
    V["現在庫数"]   = cur
    V["お気に入り"] = (st.fav if st and st.fav is not None else 0)

    # 最終入荷日（除外3条件・確定№12：空白/先頭_/-SAI- を除外）R31
    arr = q1("""SELECT MAX(arrival_date) d FROM `analytics_layer.inventory_snapshot`
      WHERE product_code=@pc AND delivery_note_no IS NOT NULL AND delivery_note_no!=''
        AND NOT STARTS_WITH(delivery_note_no,'_')
        AND NOT REGEXP_CONTAINS(delivery_note_no,'-SAI-')""", pc=("STRING", pc))
    V["最終入荷日"] = asdate(arr.d) if arr else None

    # 予約未処理数 R50
    rv = q1("""SELECT SUM(quantity) q FROM `analytics_layer.reservations` WHERE product_code=@pc
      AND reservation_date=(SELECT MAX(reservation_date) FROM `analytics_layer.reservations`
                            WHERE product_code=@pc)""", pc=("STRING", pc))
    rp = (rv.q if rv and rv.q is not None else 0)
    V["予約未処理数"] = rp

    # 入荷残（フリー在庫用）＋入荷山1/2/3 R52-59
    inc = q1("""SELECT SUM(incoming_qty) q FROM `analytics_layer.incoming_stock` WHERE product_code=@pc
      AND source_date=(SELECT MAX(source_date) FROM `analytics_layer.incoming_stock`
                       WHERE product_code=@pc)""", pc=("STRING", pc))
    remain = (inc.q if inc and inc.q is not None else 0)
    V["_arrivals"] = qn("""
      SELECT SAFE_CAST(REPLACE(earliest_arrival_date,'/','-') AS DATE) d, SUM(incoming_qty) q
      FROM `analytics_layer.incoming_stock`
      WHERE product_code=@pc AND earliest_arrival_date IS NOT NULL
        AND source_date=(SELECT MAX(source_date) FROM `analytics_layer.incoming_stock` WHERE product_code=@pc)
      GROUP BY d HAVING d IS NOT NULL AND d>=@asof ORDER BY d LIMIT 3""",
      pc=("STRING", pc), asof=("DATE", asof))

    # 累計レビュー件数/点数（作成日時点・最新累計）R12-13
    rr = q1("""SELECT COUNT(*) c, ROUND(AVG(rating),2) a FROM `analytics_layer.product_reviews`
      WHERE product_code=@pc AND review_date<=@asof""", pc=("STRING", pc), asof=("DATE", asof))
    V["累計レビュー件数"] = rr.c if rr else 0
    V["累計レビュー点数"] = rr.a if rr else None

    # 日次系列 → 日販平均(短期)・中央値(長期, 在庫ルール込み) R38/R43
    dsl = {r.d.isoformat(): r.q for r in qn("""SELECT sale_date d, SUM(sales_quantity) q
      FROM `analytics_layer.sales_daily` WHERE product_code=@pc AND source_file='orders'
      AND sale_date BETWEEN @dl AND @asof GROUP BY d""",
      pc=("STRING", pc), dl=("DATE", dL), asof=("DATE", asof))}
    dst = {r.d.isoformat(): r.s for r in qn("""SELECT snapshot_date d, SUM(available_qty) s
      FROM `analytics_layer.stock_analysis` WHERE product_code=@pc
      AND snapshot_date BETWEEN @dl AND @asof GROUP BY d""",
      pc=("STRING", pc), dl=("DATE", dL), asof=("DATE", asof))}

    def dval(ds):  # R43: 受注>0→値, 在庫>0&受注0→0, それ以外→除外
        qd = dsl.get(ds, 0) or 0
        if qd > 0: return qd
        s = dst.get(ds)
        return 0 if (s and s > 0) else None

    daysS = [(asof_d - datetime.timedelta(days=i)).isoformat() for i in range(WIN_SHORT)]
    daysL = [(asof_d - datetime.timedelta(days=i)).isoformat() for i in range(WIN_LONG)]
    sumS  = sum((dsl.get(d, 0) or 0) for d in daysS)
    elapsed = WIN_SHORT
    if V["販売開始日"]:
        elapsed = min(WIN_SHORT, (asof_d - datetime.date.fromisoformat(V["販売開始日"])).days + 1)
    elapsed = max(1, elapsed)
    veloS = round(sumS / elapsed, 2)                                           # R38
    valsL = [v for v in (dval(d) for d in daysL) if v is not None]
    median = round(statistics.median(valsL), 2) if valsL else 0               # R43
    V[f"{WIN_SHORT}日販売数"] = sumS
    V[f"{WIN_SHORT}日日販平均"] = veloS
    V[f"{WIN_LONG}日販売数"] = sum((dsl.get(d, 0) or 0) for d in daysL)
    V[f"{WIN_LONG}日日販中央値"] = median

    # 在庫日数・完売想定日（分母は CONFIG で切替可能）R39/40/44/45/48
    basis = median if STOCK_DAYS_BASIS == "median" else veloS
    def sdays(stock, rate): return round(stock / rate, 1) if rate else None
    def soldout(days): return (asof_d + datetime.timedelta(days=round(days))).isoformat() if days is not None else None
    sdS = sdays(cur, veloS)
    sdL = sdays(cur, median)
    V[f"{WIN_SHORT}日現在庫日数"]   = sdS
    V[f"{WIN_SHORT}日完売想定日"]   = soldout(sdS)
    V[f"{WIN_LONG}日現在庫日数"]    = sdL
    V[f"{WIN_LONG}日完売想定日"]    = soldout(sdL)
    free = cur + remain - rp                                                   # フリー在庫 確定№3
    V["フリー在庫数"] = free
    V["フリー在庫日数"] = sdays(free, basis)

    # ② 内訳: UU（sales_daily由来）/ 時系列ピボット R65/R80-82
    uu = q1("SELECT SUM(unique_visitors) u FROM `analytics_layer.sales_daily` "
            "WHERE product_code=@pc AND sale_date BETWEEN @sd AND @ed",
            pc=("STRING", pc), sd=("DATE", start), ed=("DATE", end))
    V["UU"] = uu.u if uu and uu.u is not None else None
    def piv(fmt):
        return qn(f"SELECT {fmt} k, SUM(sales_quantity) q FROM `analytics_layer.sales_daily` "
                  "WHERE product_code=@pc AND source_file='orders' AND sale_date BETWEEN @sd AND @ed "
                  "GROUP BY k ORDER BY k", pc=("STRING", pc), sd=("DATE", start), ed=("DATE", end))
    V["_piv_y"] = [(str(int(r.k)), r.q) for r in piv("EXTRACT(YEAR FROM sale_date)")]
    V["_piv_m"] = [(r.k, r.q) for r in piv("FORMAT_DATE('%Y/%m', sale_date)")]
    V["_piv_d"] = [(r.k, r.q) for r in piv("FORMAT_DATE('%Y/%m/%d', sale_date)")]
    return V


# ============================ レイアウト定義（仕様変更はここ）============================
# (種別, ラベル, 値キー or 固定値, 注記)   種別: sec=見出し / item=項目 / blank=空行
SHORT, LONG = WIN_SHORT, WIN_LONG
REPORT_LAYOUT = [
    ("sec",  "基本情報", None, ""),
    ("item", "品番", "品番", ""),
    ("item", "商品名", "商品名", ""),
    ("item", "ショップ", "ショップ", ""),
    ("item", "集計開始日", "集計開始日", "入力値"),
    ("item", "集計終了日", "集計終了日", "入力値"),
    ("item", "作成日", "作成日", "自動（データ最新日）"),
    ("blank", "", None, ""),
    ("sec",  "① 集計", None, ""),
    ("item", "ブランド（親カテゴリ）", "ブランド", "年2回【限定】回避は直近の正ブランド引継ぎ"),
    ("item", "商品タイプ親", "商品タイプ親", ""),
    ("item", "商品タイプ子", "商品タイプ子", ""),
    ("item", "上代（税抜）", "上代", ""),
    ("item", "販売開始日", "販売開始日", "期間内初回受注日（確定№2）"),
    ("item", "累計レビュー件数", "累計レビュー件数", ""),
    ("item", "累計レビュー点数", "累計レビュー点数", ""),
    ("item", "合計販売数", "合計販売数", ""),
    ("item", "平均売価（税抜）", "平均売価", ""),
    ("item", "合計値引率(%)", "値引率", ""),
    ("item", "合計粗利率(%)", "粗利率", "PF→MMS（R14/R69）"),
    ("item", "最新加重平均原価", "最新加重平均原価", ""),
    ("item", "お気に入り", "お気に入り", "在庫分析daily（確定№11）"),
    ("item", "現在庫数", "現在庫数", "S列販売可能数・予約0（確定№1）"),
    ("item", "最終入荷日", "最終入荷日", "納品書NO 空白/先頭_/-SAI- 除外（確定№12）"),
    ("item", "予約未処理数", "予約未処理数", ""),
    ("item", "フリー在庫数", "フリー在庫数", "現在庫(予約0)+入荷残-予約未処理（確定№3）"),
    ("item", "フリー在庫日数", "フリー在庫日数", f"分母={STOCK_DAYS_BASIS}（要客確認の矛盾点）"),
    ("item", "前回発注日", None, "MMS発注書一覧/sitateru 未取得"),
    ("item", "前回原価", None, "予約管理表/sitateru 未取得"),
    ("item", "画像", None, "カラーコード(データ連携管理)未取得"),
    ("item", "確定発注数", "", "手入力欄"),
    ("blank", "", None, ""),
    ("item", f"▼直近{SHORT}日 販売数", f"{SHORT}日販売数", ""),
    ("item", f"▼直近{SHORT}日 日販平均", f"{SHORT}日日販平均", "販売数÷7（7日未満は経過日数）"),
    ("item", f"▼直近{SHORT}日 現在庫日数", f"{SHORT}日現在庫日数", ""),
    ("item", f"▼直近{SHORT}日 完売想定日", f"{SHORT}日完売想定日", ""),
    ("item", f"▼直近{LONG}日 販売数", f"{LONG}日販売数", ""),
    ("item", f"▼直近{LONG}日 日販中央値", f"{LONG}日日販中央値", "在庫あり受注0=0/在庫無し受注0=除外"),
    ("item", f"▼直近{LONG}日 現在庫日数", f"{LONG}日現在庫日数", ""),
    ("item", f"▼直近{LONG}日 完売想定日", f"{LONG}日完売想定日", ""),
    ("blank", "", None, ""),
    ("group_arrivals", "入荷残（予約管理表）", "_arrivals", ""),
    ("group_fku", "FKU枚数構成（品番&カラー）", "_fku", ""),
    ("blank", "", None, ""),
    ("sec",  "② 内訳", None, ""),
    ("item", "合計販売数", "合計販売数", ""),
    ("item", "平均売価（税抜）", "平均売価", ""),
    ("item", "粗利率(%)", "粗利率", ""),
    ("item", "値引率(%)", "値引率", ""),
    ("item", "予約販売数割合(%)", "予約販売数割合", ""),
    ("item", "UU", "UU", "sales_daily由来・CVRと1日ずれ"),
    ("item", "CVR(%)", None, "受注点数(ダッシュボード値)未配線"),
    ("item", "お気に率(%)", None, "お気に入り登録者数(ダッシュボード値)未配線"),
    ("item", "CP対象枚数比(%)", None, "クーポン実施日判定 要確認"),
    ("item", "入荷数量（期間）", None, "MMS発注書一覧 未取得"),
    ("blank", "", None, ""),
    ("group_piv", "時系列（注文数）年", "_piv_y", ""),
    ("group_piv", "時系列（注文数）月", "_piv_m", ""),
    ("group_piv", "時系列（注文数）日", "_piv_d", ""),
]


def build_rows(V):
    """REPORT_LAYOUT と V から [種別, ラベル, 値, 注記] の2D行を生成（出力先共通）。"""
    out = [["title", "発注管理表（項目詳細） — 品番×期間の実数", "", ""]]
    for kind, label, key, note in REPORT_LAYOUT:
        if kind == "sec":
            out.append(["sec", label, "", ""])
        elif kind == "blank":
            out.append(["blank", "", "", ""])
        elif kind == "item":
            if key is None:
                val = NA
            elif key in V:
                val = V[key]
                val = NA if val is None else val
            else:
                val = key  # 固定値（例: 確定発注数=""）
            out.append(["item", label, "" if val is None else val, note])
        elif kind == "group_arrivals":
            out.append(["sec", label, "", ""])
            ar = V.get(key, [])
            for i in range(3):
                if i < len(ar):
                    out.append(["item", f"入荷日{i+1}", asdate(ar[i].d), ""])
                    out.append(["item", f"　入荷数{i+1}", ar[i].q, ""])
                else:
                    out.append(["item", f"入荷日{i+1}", "", ""])
                    out.append(["item", f"　入荷数{i+1}", "", ""])
        elif kind == "group_fku":
            out.append(["sec", label, "", ""])
            for col, (pct, n) in sorted(V.get(key, {}).items(), key=lambda x: -x[1][0]):
                out.append(["item", f"　{col}", f"{pct}%", f"{n}枚"])
        elif kind == "group_piv":
            out.append(["sec", label, "", ""])
            for k, qv in V.get(key, []):
                out.append(["item", f"　{k}", qv, ""])
    return out


# ============================ 出力先（差し替え可能）============================
def render_xlsx(rows, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook(); ws = wb.active; ws.title = "発注管理表"
    f_title = Font(bold=True, color="FFFFFF"); fill_t = PatternFill("solid", fgColor="374151")
    f_sec = Font(bold=True); fill_s = PatternFill("solid", fgColor="E5E7EB")
    f_note = Font(italic=True, color="9CA3AF", size=9)
    for i, (kind, label, val, note) in enumerate(rows, start=1):
        ws.cell(i, 1, label); ws.cell(i, 2, val); ws.cell(i, 3, note).font = f_note
        if kind == "title":
            for c in (1, 2, 3): ws.cell(i, c).fill = fill_t
            ws.cell(i, 1).font = f_title
        elif kind == "sec":
            for c in (1, 2, 3): ws.cell(i, c).fill = fill_s
            ws.cell(i, 1).font = f_sec
    ws.column_dimensions["A"].width = 26; ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 46
    wb.save(path); return path

def render_gsheet(rows, sheet_id, tab="発注管理表"):
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build
    creds = Credentials.from_service_account_file(
        SA_KEY, scopes=["https://www.googleapis.com/auth/spreadsheets"])
    svc = build("sheets", "v4", credentials=creds)
    meta = svc.spreadsheets().get(spreadsheetId=sheet_id).execute()
    tabs = [s["properties"]["title"] for s in meta["sheets"]]
    if tab not in tabs:
        svc.spreadsheets().batchUpdate(spreadsheetId=sheet_id,
            body={"requests": [{"addSheet": {"properties": {"title": tab}}}]}).execute()
    svc.spreadsheets().values().clear(spreadsheetId=sheet_id, range=tab).execute()
    values = [[label, ("" if val is None else val), note] for _, label, val, note in rows]
    svc.spreadsheets().values().update(spreadsheetId=sheet_id, range=f"{tab}!A1",
        valueInputOption="RAW", body={"values": values}).execute()
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/edit#gid=0"


def main():
    pc    = sys.argv[1] if len(sys.argv) > 1 else os.getenv("PC", "sc1032")
    start = sys.argv[2] if len(sys.argv) > 2 else os.getenv("START", "2026-05-01")
    end   = sys.argv[3] if len(sys.argv) > 3 else os.getenv("END", "2026-05-31")
    V = compute_values(pc, start, end)
    rows = build_rows(V)

    os.makedirs(OUT_DIR, exist_ok=True)
    xpath = os.path.join(OUT_DIR, f"発注管理表_{pc}_{start}_{end}.xlsx")
    render_xlsx(rows, xpath)
    print(f"OK xlsx: {xpath}")

    gid = os.getenv("GSHEET_ID")
    if gid:
        try:
            url = render_gsheet(rows, gid)
            print(f"OK スプシ: {url}")
        except Exception as e:
            print(f"スプシ書込失敗（共有設定を確認 / SA={SA_KEY}）: {str(e)[:200]}")
    else:
        print("（GSHEET_ID未指定 → スプシ出力はスキップ。共有済みシートIDを GSHEET_ID に指定すると書込）")

    print(f"  品番={pc} 期間={start}〜{end} 作成日={V['作成日']}")
    print(f"  販売数={V['合計販売数']} 平均売価={V['平均売価']} 値引率={V['値引率']}% 粗利率={V['粗利率']}%")


if __name__ == "__main__":
    main()
