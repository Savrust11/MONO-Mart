"""
クライアント提供の過去データ (受注・発送) を Google Drive → GCS へ移送する。

対象 (クライアント 古城様 2026-06-17 提供、検証済み構造):
  受注 folder (1eRZoby0Bo8J4c6anWoe0PSGpDaGjePaR)
    └─ 月フォルダ 202407..202508 (14ヶ月)
         └─ 日別 .xlsx (例: 2024_07_01.xlsx)  ← ヘッダー1行目・通常レイアウト
  発送 folder (1kSuIgbkOtpfOu474A-rFRWKcLN0wKjJK)
    └─ 月次ファイル 20240701(出荷).xlsx / 20240801(出荷).csv ... (14ヶ月)
         ← 売上集計メニューの代替ファイル。出荷日(R列)・出荷数(P列)・項目少なめ。
            1ファイルに1ヶ月分の全出荷行が入る (8月で約14.7万行)。

GCS 出力レイアウト (既存の日次データと同じ prefix。日付が 2024-2025 なので衝突なし):
  uploads/zozo/orders/{YYYY-MM-DD}/{YYYY_MM_DD}.csv     (受注: 日別)
  uploads/zozo/shipped/{YYYY-MM-DD}/{原ファイル名}.csv   (発送: 月初日フォルダ)

xlsx は CSV(UTF-8 BOM) に変換してアップロード。元から csv のものはそのまま。
列名の正規化や出荷日のパースは取り込み(ingest)時に zozo_csv_extractor が行うため、
ここでは「素のCSV化」だけを行う。

使い方:
  # テスト: 2024年7月だけ (受注+発送)
  python backfill_drive_to_gcs.py --month 202407
  # 受注の特定月の先頭3日だけ (動作確認用)
  python backfill_drive_to_gcs.py --only orders --month 202407 --limit-days 3
  # 全期間 (14ヶ月 × 受注+発送)
  python backfill_drive_to_gcs.py

ENV:
  GCS_RAW_BUCKET   (default: mono-back-office-system-raw-data)
  GOOGLE_APPLICATION_CREDENTIALS  (SA key — Drive read + GCS write 両方必要)
"""
from __future__ import annotations

import argparse
import io
import logging
import os
import re
import sys
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from google.cloud import storage
from openpyxl import load_workbook
import csv as _csv

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("backfill_drive_to_gcs")

_KEY = Path(__file__).resolve().parents[1] / "sheets-sa-key.json"
_DRIVE_SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]

ORDERS_FOLDER_ID  = "1eRZoby0Bo8J4c6anWoe0PSGpDaGjePaR"
SHIPPED_FOLDER_ID = "1kSuIgbkOtpfOu474A-rFRWKcLN0wKjJK"

BUCKET = os.getenv("GCS_RAW_BUCKET", "mono-back-office-system-raw-data")
SKIP_EXISTING = os.getenv("SKIP_EXISTING", "1") == "1"
WORKERS = int(os.getenv("BACKFILL_WORKERS", "8"))

# httplib2 (Drive service の下回り) はスレッド非安全なので、スレッドごとに
# 専用の Drive service を持つ。Credentials は共有してよい。
_creds = Credentials.from_service_account_file(str(_KEY), scopes=_DRIVE_SCOPES)
_tlocal = threading.local()


def _drive():
    """スレッドローカルな Drive service を返す (並列実行で安全)。"""
    if not hasattr(_tlocal, "drive"):
        _tlocal.drive = build("drive", "v3", credentials=_creds,
                              cache_discovery=False)
    return _tlocal.drive


def _list(folder_id: str) -> list[dict]:
    """1階層分のファイル/フォルダを名前順で返す (共有ドライブ対応)。"""
    drive = _drive()
    out: list[dict] = []
    token = None
    while True:
        res = drive.files().list(
            q=f"'{folder_id}' in parents and trashed=false",
            fields="nextPageToken, files(id,name,mimeType,size)",
            supportsAllDrives=True, includeItemsFromAllDrives=True,
            orderBy="name", pageSize=1000, pageToken=token).execute()
        out.extend(res.get("files", []))
        token = res.get("nextPageToken")
        if not token:
            break
    return out


def _is_folder(f: dict) -> bool:
    return f["mimeType"].endswith("folder")


def _download(file_id: str) -> bytes:
    drive = _drive()
    buf = io.BytesIO()
    dl = MediaIoBaseDownload(buf, drive.files().get_media(
        fileId=file_id, supportsAllDrives=True))
    done = False
    while not done:
        _, done = dl.next_chunk()
    return buf.getvalue()


def _existing_keys(gcs: storage.Client, prefix: str) -> set[str]:
    """GCS に既にあるキー集合 (resume 用)。"""
    return {b.name for b in gcs.list_blobs(BUCKET, prefix=prefix)}


def _cell(v) -> str:
    """CSVセル文字列化。日付/数値は後段(zozo_csv_extractor)が解釈できる形にする。"""
    if v is None or v == "":
        return ""
    # datetime/date → "YYYY-MM-DD ..." (_normalize_date が先頭の日付部分を解釈)
    return str(v)


def _xlsx_to_csv(data: bytes) -> bytes:
    """xlsx の先頭シートを CSV(UTF-8 BOM) に変換。
    python-calamine (Rust製・高速) を優先。失敗時は openpyxl にフォールバック。"""
    try:
        from python_calamine import CalamineWorkbook
        wb = CalamineWorkbook.from_filelike(io.BytesIO(data))
        rows = wb.get_sheet_by_index(0).to_python()
        out = io.StringIO()
        w = _csv.writer(out)
        for row in rows:
            w.writerow([_cell(c) for c in row])
        return out.getvalue().encode("utf-8-sig")
    except Exception as exc:
        logger.warning("calamine変換失敗→openpyxlへ: %s", exc)
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        out = io.StringIO()
        w = _csv.writer(out)
        for row in ws.iter_rows(values_only=True):
            w.writerow(["" if c is None else c for c in row])
        wb.close()
        return out.getvalue().encode("utf-8-sig")


def _date_from_name(name: str) -> str | None:
    """ファイル名から YYYY-MM-DD を抽出 (最初の8桁数字)。
    '2024_07_01.xlsx' → 2024-07-01 / '20240801(出荷).csv' → 2024-08-01。"""
    digits = re.sub(r"\D", "", name)
    if len(digits) >= 8:
        y, m, d = digits[0:4], digits[4:6], digits[6:8]
        return f"{y}-{m}-{d}"
    return None


def _upload(gcs: storage.Client, key: str, data: bytes) -> None:
    gcs.bucket(BUCKET).blob(key).upload_from_string(
        data, content_type="text/csv; charset=utf-8")


def _to_csv_bytes(f: dict) -> bytes | None:
    """Drive ファイルを CSV bytes にして返す (xlsx は変換、csv はそのまま)。"""
    raw = _download(f["id"])
    name = f["name"].lower()
    if name.endswith(".xlsx") or "spreadsheetml" in f["mimeType"]:
        try:
            return _xlsx_to_csv(raw)
        except Exception as exc:
            logger.error("  xlsx変換失敗 %s: %s", f["name"], exc)
            return None
    return raw  # 既に csv


def _process_one(gcs: storage.Client, f: dict, key: str) -> tuple[str, str]:
    """1ファイルを download→変換→upload。worker スレッドで実行。
    return (status, message)。status: ok / skip / fail。"""
    try:
        csv_bytes = _to_csv_bytes(f)
        if csv_bytes is None:
            return ("fail", f"{f['name']}: 変換失敗")
        _upload(gcs, key, csv_bytes)
        return ("ok", f"{f['name']} → {key} ({len(csv_bytes):,} bytes)")
    except Exception as exc:
        return ("fail", f"{f['name']}: {type(exc).__name__}: {str(exc)[:120]}")


def _run_pool(gcs: storage.Client, tasks: list[tuple[dict, str]],
              label: str) -> int:
    """tasks = [(file_dict, gcs_key), ...] を並列処理。アップロード成功数を返す。"""
    if not tasks:
        logger.info("%s: 対象ファイルなし (全てスキップ済み)", label)
        return 0
    ok = fail = 0
    done = 0
    total = len(tasks)
    with ThreadPoolExecutor(max_workers=WORKERS) as ex:
        futs = {ex.submit(_process_one, gcs, f, key): f["name"]
                for f, key in tasks}
        for fut in as_completed(futs):
            status, msg = fut.result()
            done += 1
            if status == "ok":
                ok += 1
                logger.info("  [%s %d/%d] ✓ %s", label, done, total, msg)
            else:
                fail += 1
                logger.error("  [%s %d/%d] ✗ %s", label, done, total, msg)
    logger.info("%s: %d 成功 / %d 失敗", label, ok, fail)
    return ok


def backfill_orders(gcs, month: str | None, limit_days: int | None) -> int:
    months = [f for f in _list(ORDERS_FOLDER_ID) if _is_folder(f)]
    if month:
        months = [m for m in months if m["name"] == month]
    logger.info("受注: 対象月 %d (%s)", len(months), [m["name"] for m in months])
    existing = _existing_keys(gcs, "uploads/zozo/orders/") if SKIP_EXISTING else set()

    tasks: list[tuple[dict, str]] = []
    for mf in months:
        days = [f for f in _list(mf["id"]) if not _is_folder(f)]
        if limit_days:
            days = days[:limit_days]
        for f in days:
            date_iso = _date_from_name(f["name"])
            if not date_iso:
                logger.warning("  日付不明 skip: %s", f["name"]); continue
            key = f"uploads/zozo/orders/{date_iso}/{date_iso.replace('-', '')}.csv"
            if key in existing:
                continue
            tasks.append((f, key))
    logger.info("受注: %d ファイルを処理 (skip済み除く)", len(tasks))
    return _run_pool(gcs, tasks, "受注")


def backfill_shipped(gcs, month: str | None) -> int:
    files = [f for f in _list(SHIPPED_FOLDER_ID) if not _is_folder(f)]
    if month:
        files = [f for f in files if re.sub(r"\D", "", f["name"])[:6] == month]
    existing = _existing_keys(gcs, "uploads/zozo/shipped/") if SKIP_EXISTING else set()

    tasks: list[tuple[dict, str]] = []
    for f in files:
        date_iso = _date_from_name(f["name"])  # 月初日 (YYYY-MM-01)
        if not date_iso:
            logger.warning("  日付不明 skip: %s", f["name"]); continue
        base = re.sub(r"\.(xlsx|csv)$", "", f["name"], flags=re.I)
        key = f"uploads/zozo/shipped/{date_iso}/{base}.csv"
        if key in existing:
            continue
        tasks.append((f, key))
    logger.info("発送: %d ファイルを処理 (skip済み除く)", len(tasks))
    return _run_pool(gcs, tasks, "発送")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--only", choices=["orders", "shipped"], default=None,
                    help="片方だけ処理 (default: 両方)")
    ap.add_argument("--month", default=None, help="YYYYMM 1ヶ月だけ (例: 202407)")
    ap.add_argument("--limit-days", type=int, default=None,
                    help="受注: 月内の先頭N日だけ (動作確認用)")
    args = ap.parse_args()

    if not _KEY.exists():
        logger.error("SA key が見つかりません: %s", _KEY)
        return 2

    gcs = storage.Client()
    logger.info("GCS bucket = %s  (workers=%d, skip_existing=%s)",
                BUCKET, WORKERS, SKIP_EXISTING)

    total = 0
    if args.only in (None, "orders"):
        total += backfill_orders(gcs, args.month, args.limit_days)
    if args.only in (None, "shipped"):
        total += backfill_shipped(gcs, args.month)

    logger.info("=== 完了: 合計 %d ファイルを GCS へ移送 ===", total)
    return 0


if __name__ == "__main__":
    sys.exit(main())
