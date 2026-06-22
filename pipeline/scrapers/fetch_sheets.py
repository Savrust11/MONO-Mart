"""
Google Sheets + Drive fetcher.

Sources:
  1. Google Sheets — PF手数料, 発注明細 (predefined spreadsheet IDs)
  2. Drive folder ZOZO買い回りデータ — every monthly xlsx in the folder
     is downloaded, converted to CSV (first worksheet) and uploaded.

Auth: pipeline/sheets-sa-key.json  (SA = sheets-fetcher@mono-back-office-system
.iam.gserviceaccount.com). The client must SHARE the parent Drive folder with
that SA email (Viewer is enough).

Env: GCS_RAW_BUCKET, TARGET_DATE (default yesterday JST).
"""
from __future__ import annotations
import os, sys, csv, io, json, re
from datetime import datetime, timedelta, timezone
from pathlib import Path

import gspread
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

KEY = Path(__file__).resolve().parents[1] / "sheets-sa-key.json"
JST = timezone(timedelta(hours=9))
SCOPES = ["https://www.googleapis.com/auth/spreadsheets.readonly",
          "https://www.googleapis.com/auth/drive.readonly"]

# (key, spreadsheet_id, tabs (None=first only, "*"=all combined), GCS subfolder, output filename)
TARGETS = [
    ("pf_fee",  "1z7b7WQpzw_gKJUbIcMWCz6zQyRCSDLKkRNVr5aTY4Xs", None,
     "sheets/pf_fee", "pf_fee.csv"),
    # 発注明細 has two tabs (プロダクト1部 / プロダクト2部) — combine both with a
    # leading '_tab' column so downstream ingest can keep them separated if needed
    ("hacchu",  "1iEyR2IzLtUUA1Eu0iTE9BUDz11CcsUoAAxPTOgjzpQk", "*",
     "tableau/hacchu", "発注明細.csv"),
]

# Drive xlsx-folder sources: every xlsx in the folder is downloaded + converted.
# (key, drive_folder_id, GCS subfolder)
XLSX_FOLDER_TARGETS = [
    ("buy_round", "1H6NJdhZD49SVbOajKmrNnpaL3Wnwp2wB", "sheets/buy_round"),
]


def main():
    if not KEY.exists() or KEY.stat().st_size < 100:
        print(f"FATAL: SA key missing/empty at {KEY}")
        return 2
    target_date = os.getenv("TARGET_DATE") or (datetime.now(JST) - timedelta(days=1)).strftime("%Y-%m-%d")
    bucket = os.getenv("GCS_RAW_BUCKET", "mono-back-office-system-raw-data")

    creds = Credentials.from_service_account_file(str(KEY), scopes=SCOPES)
    sa_email = json.loads(KEY.read_text(encoding="utf-8")).get("client_email")
    print(f"SA: {sa_email}")
    gc = gspread.authorize(creds)

    results, ok_any = [], False
    for key, sid, tab, sub, fname in TARGETS:
        try:
            sh = gc.open_by_key(sid)
            if tab == "*":
                rows = []
                header = None
                for w in sh.worksheets():
                    vals = w.get_all_values()
                    if not vals:
                        continue
                    if header is None:
                        header = ["_tab"] + vals[0]
                        rows.append(header)
                    for r in vals[1:]:
                        rows.append([w.title] + r)
            else:
                ws = sh.sheet1
                rows = ws.get_all_values()
            if not rows:
                results.append({"key": key, "status": "empty"})
                print(f"[{key}] sheet empty")
                continue
            buf = io.StringIO()
            csv.writer(buf).writerows(rows)
            data = buf.getvalue().encode("utf-8-sig")  # Excel-friendly BOM
            from google.cloud import storage
            gp = f"uploads/{sub}/{target_date}/{fname}"
            storage.Client().bucket(bucket).blob(gp).upload_from_string(
                data, content_type="text/csv")
            print(f"[{key}] {len(rows)} rows → gs://{bucket}/{gp}")
            results.append({"key": key, "status": "ok", "rows": len(rows),
                            "gcs": f"gs://{bucket}/{gp}"})
            ok_any = True
        except gspread.exceptions.APIError as e:
            msg = str(e)[:160]
            hint = (" — SHARE the sheet with the SA email above (Viewer)"
                    if "PERMISSION" in msg.upper() or "403" in msg else "")
            print(f"[{key}] APIError: {msg}{hint}")
            results.append({"key": key, "status": "permission_or_api", "err": msg})
        except Exception as e:
            print(f"[{key}] FAIL {type(e).__name__}: {str(e)[:160]}")
            results.append({"key": key, "status": "fail", "err": str(e)[:160]})

    # --- Drive xlsx folders (ブランド別買い回りデータ etc.) -----------------
    drive = build("drive", "v3", credentials=creds, cache_discovery=False)
    from google.cloud import storage
    storage_client = storage.Client()

    for key, folder_id, sub in XLSX_FOLDER_TARGETS:
        try:
            res = drive.files().list(
                q=(f"'{folder_id}' in parents and trashed=false "
                   f"and (mimeType='application/vnd.openxmlformats-"
                   f"officedocument.spreadsheetml.sheet' "
                   f"or mimeType='application/vnd.ms-excel')"),
                fields="files(id,name,mimeType,modifiedTime)",
                supportsAllDrives=True,
                includeItemsFromAllDrives=True,
                orderBy="name desc",
                pageSize=50).execute()
            files = res.get("files", [])
            if not files:
                results.append({"key": key, "status": "empty_folder"})
                print(f"[{key}] folder has no xlsx files")
                continue
            # Each monthly file is ~90 MB after CSV conversion; loading all 11
            # would OOM the worker. Pick the latest 2 months only — enough for
            # rolling MoM comparison.
            files = files[:2]
            sub_ok = 0
            for f in files:
                buf = io.BytesIO()
                req = drive.files().get_media(fileId=f["id"],
                                              supportsAllDrives=True)
                downloader = MediaIoBaseDownload(buf, req)
                done = False
                while not done:
                    _, done = downloader.next_chunk()
                xlsx_bytes = buf.getvalue()
                csv_bytes = _xlsx_to_csv(xlsx_bytes)
                if csv_bytes is None:
                    print(f"[{key}] {f['name']} → openpyxl conversion failed")
                    continue
                base = Path(f["name"]).stem
                gp = f"uploads/{sub}/{target_date}/{base}.csv"
                storage_client.bucket(bucket).blob(gp).upload_from_string(
                    csv_bytes, content_type="text/csv")
                print(f"[{key}] {f['name']} → gs://{bucket}/{gp} "
                      f"({len(csv_bytes):,} bytes)")
                sub_ok += 1
            results.append({"key": key, "status": "ok", "files": sub_ok})
            ok_any = True
        except Exception as e:
            print(f"[{key}] FAIL {type(e).__name__}: {str(e)[:160]}")
            results.append({"key": key, "status": "fail", "err": str(e)[:160]})

    print("RESULTS:", json.dumps(results, ensure_ascii=False))
    return 0 if ok_any else 1


def _xlsx_to_csv(data: bytes) -> bytes | None:
    """Convert first worksheet of an xlsx blob to UTF-8 CSV bytes."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("openpyxl missing — run: pip install openpyxl")
        return None
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = wb.active
    buf = io.StringIO()
    w = csv.writer(buf)
    for row in ws.iter_rows(values_only=True):
        w.writerow(["" if c is None else c for c in row])
    return buf.getvalue().encode("utf-8-sig")


if __name__ == "__main__":
    sys.exit(main())
