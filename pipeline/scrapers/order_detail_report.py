# -*- coding: utf-8 -*-
"""
発注管理表（項目詳細）— 品番 × 指定期間 の「実数」レポート生成。
顧客最優先機能：任意の品番・指定期間を渡すと、①集計・②内訳の実数がスプシ(.xlsx)で出る。

仕様：シート「発注管理表項目詳細」(R02-R82) の計算式どおりに実装（独自計算なし）。
確定事項（社内で決めること）反映：
  ・出力＝スプシ
  ・現在庫数＝S列販売可能数(available_qty)、販売タイプ予約は0
  ・フリー在庫＝現在庫(予約0)＋入荷残−予約未処理
  ・最終入荷日除外＝納品書NO空白/先頭"_"/"-SAI-"含む を除外
  ・お気に入り＝在庫分析daily
  ・販売開始日＝期間内初回受注日
データ未取得の項目（前回発注日/原価・画像カラーコード・CP対象枚数比）は空欄＋注記。
在庫日数の分母は 30日中央値 に統一（フリー在庫日数の"平均"記載は仕様内矛盾のため。要客確認）。

使い方:
  python order_detail_report.py <品番> <開始日 YYYY-MM-DD> <終了日 YYYY-MM-DD>
"""
import os, sys, datetime, statistics
from google.cloud import bigquery
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

PROJECT = "mono-back-office-system"
bq = bigquery.Client(project=PROJECT)

PC    = sys.argv[1] if len(sys.argv) > 1 else os.getenv("PC", "sc1032")
START = sys.argv[2] if len(sys.argv) > 2 else os.getenv("START", "2026-05-01")
END   = sys.argv[3] if len(sys.argv) > 3 else os.getenv("END", "2026-05-31")

# 作成日（自動）。直近7日/30日は作成日基準。データ最新日を超えないようにする。
TODAY = datetime.date.today().isoformat()
ASOF = list(bq.query(
    "SELECT LEAST(@t, CAST(MAX(sale_date) AS STRING)) d FROM `analytics_layer.sales_daily` "
    "WHERE source_file='orders'",
    job_config=bigquery.QueryJobConfig(query_parameters=[
        bigquery.ScalarQueryParameter("t", "STRING", TODAY)]),
).result())[0].d
asof_d = datetime.date.fromisoformat(ASOF)
d7  = (asof_d - datetime.timedelta(days=6)).isoformat()
d30 = (asof_d - datetime.timedelta(days=29)).isoformat()

def P(**kw):
    m = {"STRING": str, "DATE": str}
    return bigquery.QueryJobConfig(query_parameters=[
        bigquery.ScalarQueryParameter(k, t, v) for k, (t, v) in kw.items()])

def one(sql, **kw):
    r = list(bq.query(sql, job_config=P(**kw)).result())
    return r[0] if r else None

def asdate(v):
    if v is None or v == "":
        return None
    return v.isoformat() if hasattr(v, "isoformat") else str(v)

def rows(sql, **kw):
    return list(bq.query(sql, job_config=P(**kw)).result())

# ---------- 1) 受注期間集計（SKU単位） + 原価（PF品番→MMS SKU） ----------
sku_rows = rows("""
WITH s AS (
  SELECT sku_code, color_name, size, sale_type,
    SUM(sales_quantity) qty, SUM(sales_amount) rev,
    SUM(proper_price*sales_quantity) lst
  FROM `analytics_layer.sales_daily`
  WHERE product_code=@pc AND source_file='orders'
    AND sale_date BETWEEN @sd AND @ed AND sales_quantity>0
  GROUP BY sku_code, color_name, size, sale_type
),
pfc AS (
  SELECT ANY_VALUE(cost_price) val FROM `analytics_layer.pf_fee_master`
  WHERE UPPER(TRIM(product_code))=UPPER(TRIM(@pc))
    AND snapshot_date=(SELECT MAX(snapshot_date) FROM `analytics_layer.pf_fee_master`)
    AND cost_price>0
),
mms AS (
  SELECT j_col, j_sz, ARRAY_AGG(cost_price ORDER BY source_date DESC LIMIT 1)[OFFSET(0)] cp
  FROM (SELECT UPPER(TRIM(color_name)) j_col, UPPER(TRIM(size)) j_sz, cost_price, source_date
        FROM `analytics_layer.cost_master`
        WHERE UPPER(TRIM(product_code))=UPPER(TRIM(@pc)) AND color_name IS NOT NULL
          AND size IS NOT NULL AND valid_to IS NULL)
  GROUP BY j_col, j_sz
)
SELECT s.sku_code, s.color_name, s.size, s.sale_type, s.qty, s.rev, s.lst,
  COALESCE((SELECT val FROM pfc), mms.cp) AS cost
FROM s LEFT JOIN mms ON mms.j_col=UPPER(TRIM(s.color_name)) AND mms.j_sz=UPPER(TRIM(s.size))
""", pc=("STRING", PC), sd=("DATE", START), ed=("DATE", END))

tot_qty = sum(r.qty for r in sku_rows) or 0
tot_rev = sum(float(r.rev) for r in sku_rows) or 0.0
tot_lst = sum(float(r.lst) for r in sku_rows) or 0.0
tot_cost = sum(float(r.cost) * r.qty for r in sku_rows if r.cost is not None)
yoyaku_qty = sum(r.qty for r in sku_rows if r.sale_type and "予約" in str(r.sale_type))
cost_missing = sum(1 for r in sku_rows if r.cost is None)

avg_price   = round(tot_rev / tot_qty, 1) if tot_qty else None
discount_pc = round((1 - tot_rev / tot_lst) * 100, 1) if tot_lst else None
margin_pc   = round((tot_rev - tot_cost) / tot_rev * 100, 1) if tot_rev else None
yoyaku_rate = round(yoyaku_qty / tot_qty * 100, 1) if tot_qty else None

# FKU枚数構成（品番&カラー単位の販売数構成比） R25
fku = {}
for r in sku_rows:
    fku[r.color_name] = fku.get(r.color_name, 0) + r.qty
fku_comp = {c: round(q / tot_qty * 100, 1) for c, q in fku.items()} if tot_qty else {}

# ---------- 2) マスタ属性（品番） R07-R11, R20 ----------
m = one("""
SELECT ANY_VALUE(product_name) nm, ANY_VALUE(shop_name) shop,
  ANY_VALUE(parent_category) brand, ANY_VALUE(parent_item_type) pit,
  ANY_VALUE(child_item_type) cit, ANY_VALUE(proper_price) jodai
FROM `analytics_layer.product_master`
WHERE UPPER(TRIM(product_code))=UPPER(TRIM(@pc))
""", pc=("STRING", PC))

# ---------- 3) 販売開始日＝期間内初回受注日（確定 №2） ----------
sd0 = one("""
SELECT MIN(sale_date) d FROM `analytics_layer.sales_daily`
WHERE product_code=@pc AND source_file='orders'
  AND sale_date BETWEEN @sd AND @ed AND sales_quantity>0
""", pc=("STRING", PC), sd=("DATE", START), ed=("DATE", END))
sale_start = sd0.d.isoformat() if sd0 and sd0.d else None

# ---------- 4) 在庫分析daily（現在）：現在庫数(S列, 予約0)・お気に入り R23,R32 ----------
st = one("""
WITH pm AS (
  SELECT UPPER(TRIM(product_code)) pc, UPPER(TRIM(sku_code)) sk, ANY_VALUE(sale_type) stype
  FROM `analytics_layer.product_master` GROUP BY pc, sk
)
SELECT
  SUM(CASE WHEN pm.stype LIKE '%予約%' THEN 0 ELSE sa.available_qty END) cur_stock,
  SUM(sa.favorites) fav
FROM `analytics_layer.stock_analysis` sa
LEFT JOIN pm ON pm.pc=UPPER(TRIM(sa.product_code)) AND pm.sk=UPPER(TRIM(sa.sku_code))
WHERE sa.product_code=@pc
  AND sa.snapshot_date=(SELECT MAX(snapshot_date) FROM `analytics_layer.stock_analysis`
                        WHERE snapshot_date<=@asof)
""", pc=("STRING", PC), asof=("DATE", ASOF))
cur_stock = (st.cur_stock if st and st.cur_stock is not None else 0)
favorites = (st.fav if st and st.fav is not None else 0)

# ---------- 5) 最終入荷日（除外3条件 確定 №12） R31 ----------
arr = one("""
SELECT MAX(arrival_date) d FROM `analytics_layer.inventory_snapshot`
WHERE product_code=@pc AND delivery_note_no IS NOT NULL AND delivery_note_no!=''
  AND NOT STARTS_WITH(delivery_note_no,'_')
  AND NOT REGEXP_CONTAINS(delivery_note_no,'-SAI-')
""", pc=("STRING", PC))
last_arrival = asdate(arr.d) if arr else None

# ---------- 6) 予約未処理数 R50 ----------
rv = one("""
SELECT SUM(quantity) q FROM `analytics_layer.reservations`
WHERE product_code=@pc
  AND reservation_date=(SELECT MAX(reservation_date) FROM `analytics_layer.reservations`
                        WHERE product_code=@pc)
""", pc=("STRING", PC))
reserved_pending = (rv.q if rv and rv.q is not None else 0)

# ---------- 7) 入荷残 + 入荷山1/2/3 R52-59 ----------
inc = one("""
SELECT SUM(incoming_qty) q FROM `analytics_layer.incoming_stock`
WHERE product_code=@pc
  AND source_date=(SELECT MAX(source_date) FROM `analytics_layer.incoming_stock`
                   WHERE product_code=@pc)
""", pc=("STRING", PC))
incoming_remain = (inc.q if inc and inc.q is not None else 0)

arrivals = rows("""
SELECT SAFE_CAST(REPLACE(earliest_arrival_date,'/','-') AS DATE) d, SUM(incoming_qty) q
FROM `analytics_layer.incoming_stock`
WHERE product_code=@pc AND earliest_arrival_date IS NOT NULL
  AND source_date=(SELECT MAX(source_date) FROM `analytics_layer.incoming_stock`
                   WHERE product_code=@pc)
GROUP BY d
HAVING d IS NOT NULL AND d>=@asof
ORDER BY d LIMIT 3
""", pc=("STRING", PC), asof=("DATE", ASOF))

# ---------- 8) 累計レビュー件数/点数 R12-13（作成日時点・最新累計） ----------
rev_ = one("""
SELECT COUNT(*) cnt, ROUND(AVG(rating),2) avg_r
FROM `analytics_layer.product_reviews`
WHERE product_code=@pc AND review_date<=@asof
""", pc=("STRING", PC), asof=("DATE", ASOF))
review_cnt = rev_.cnt if rev_ else 0
review_avg = rev_.avg_r if rev_ else None

# ---------- 9) 日次系列（在庫ルール込み 日販7日平均 / 30日中央値） R38,R43 ----------
daily_sales = {r.d.isoformat(): r.q for r in rows("""
SELECT sale_date d, SUM(sales_quantity) q FROM `analytics_layer.sales_daily`
WHERE product_code=@pc AND source_file='orders' AND sale_date BETWEEN @d30 AND @asof
GROUP BY d""", pc=("STRING", PC), d30=("DATE", d30), asof=("DATE", ASOF))}
daily_stock = {r.d.isoformat(): r.s for r in rows("""
SELECT snapshot_date d, SUM(available_qty) s FROM `analytics_layer.stock_analysis`
WHERE product_code=@pc AND snapshot_date BETWEEN @d30 AND @asof
GROUP BY d""", pc=("STRING", PC), d30=("DATE", d30), asof=("DATE", ASOF))}

def day_value(dstr):
    """仕様R43: 受注>0→その値, 在庫>0&受注0→0, 在庫0/不明&受注0→除外(None)。"""
    q = daily_sales.get(dstr, 0) or 0
    if q > 0:
        return q
    s = daily_stock.get(dstr)
    if s and s > 0:
        return 0
    return None  # 除外

# 7日日販平均：販売数合計÷7（7日未満は経過日数）R38
days7 = [(asof_d - datetime.timedelta(days=i)).isoformat() for i in range(7)]
sum7 = sum((daily_sales.get(d, 0) or 0) for d in days7)
elapsed7 = min(7, (asof_d - datetime.date.fromisoformat(sale_start)).days + 1) if sale_start else 7
elapsed7 = max(1, elapsed7)
velo7 = round(sum7 / elapsed7, 2)

# 30日中央値 R43
days30 = [(asof_d - datetime.timedelta(days=i)).isoformat() for i in range(30)]
vals30 = [v for v in (day_value(d) for d in days30) if v is not None]
median30 = round(statistics.median(vals30), 2) if vals30 else 0

sales7  = sum7
sales30 = sum((daily_sales.get(d, 0) or 0) for d in days30)

# 在庫日数・完売想定日（分母は中央値に統一） R39,R40,R44,R45,R48
def stock_days(stock, rate):
    return round(stock / rate, 1) if rate else None
def soldout(days):
    return (asof_d + datetime.timedelta(days=round(days))).isoformat() if days is not None else None

sd7  = stock_days(cur_stock, velo7)
sd30 = stock_days(cur_stock, median30)
# フリー在庫（確定 №3）
free_stock = cur_stock + incoming_remain - reserved_pending
free_days  = stock_days(free_stock, median30)

# ---------- 10) ②内訳 時系列ピボット（注文数 年/月/日） R80-82 ----------
piv_y = rows("SELECT EXTRACT(YEAR FROM sale_date) k, SUM(sales_quantity) q FROM `analytics_layer.sales_daily` WHERE product_code=@pc AND source_file='orders' AND sale_date BETWEEN @sd AND @ed GROUP BY k ORDER BY k", pc=("STRING", PC), sd=("DATE", START), ed=("DATE", END))
piv_m = rows("SELECT FORMAT_DATE('%Y/%m', sale_date) k, SUM(sales_quantity) q FROM `analytics_layer.sales_daily` WHERE product_code=@pc AND source_file='orders' AND sale_date BETWEEN @sd AND @ed GROUP BY k ORDER BY k", pc=("STRING", PC), sd=("DATE", START), ed=("DATE", END))
piv_d = rows("SELECT FORMAT_DATE('%Y/%m/%d', sale_date) k, SUM(sales_quantity) q FROM `analytics_layer.sales_daily` WHERE product_code=@pc AND source_file='orders' AND sale_date BETWEEN @sd AND @ed GROUP BY k ORDER BY k", pc=("STRING", PC), sd=("DATE", START), ed=("DATE", END))

# UU（②内訳 R65）：sales_daily.unique_visitors（商品別実績由来）期間合計
uu = one("SELECT SUM(unique_visitors) u FROM `analytics_layer.sales_daily` WHERE product_code=@pc AND sale_date BETWEEN @sd AND @ed", pc=("STRING", PC), sd=("DATE", START), ed=("DATE", END))
UU = uu.u if uu and uu.u is not None else None

# ================= 出力（スプシ .xlsx）=================
NA = "（データ未取得）"
wb = Workbook(); ws = wb.active; ws.title = "発注管理表"
hdr = Font(bold=True, color="FFFFFF"); hdrfill = PatternFill("solid", fgColor="374151")
sec = Font(bold=True); secfill = PatternFill("solid", fgColor="E5E7EB")
note = Font(italic=True, color="9CA3AF", size=9)
r = 1
def put(label, value, memo=""):
    global r
    ws.cell(r, 1, label)
    ws.cell(r, 2, value if value is not None else "")
    if memo: ws.cell(r, 3, memo).font = note
    r += 1
def section(t):
    global r
    c = ws.cell(r, 1, t); c.font = sec
    for col in (1, 2, 3): ws.cell(r, col).fill = secfill
    r += 1

# タイトル
c = ws.cell(r, 1, "発注管理表（項目詳細） — 品番×期間の実数")
c.font = hdr;
for col in (1,2,3): ws.cell(r, col).fill = hdrfill
r += 1
put("品番", PC); put("商品名", m.nm if m else None); put("ショップ", m.shop if m else None)
put("集計開始日", START); put("集計終了日", END); put("作成日", ASOF, "自動（データ最新日）")
r += 1

section("① 集計")
put("ブランド（親カテゴリ）", m.brand if m else None, "年2回【限定】回避は直近の正ブランド引継ぎ")
put("商品タイプ親", m.pit if m else None)
put("商品タイプ子", m.cit if m else None)
put("上代（税抜）", float(m.jodai) if m and m.jodai else None)
put("販売開始日", sale_start, "期間内初回受注日（確定№2）")
put("累計レビュー件数", review_cnt)
put("累計レビュー点数", review_avg)
put("合計販売数", tot_qty)
put("平均売価（税抜）", avg_price)
put("合計値引率(%)", discount_pc)
put("合計粗利率(%)", margin_pc, f"PF→MMS / 原価欠落SKU={cost_missing}")
put("最新加重平均原価", round(tot_cost / tot_qty, 1) if tot_qty and tot_cost else None)
put("お気に入り", favorites, "在庫分析daily（確定№11）")
put("現在庫数", cur_stock, "S列販売可能数・予約は0（確定№1）")
put("最終入荷日", last_arrival, "納品書NO 空白/先頭_/-SAI- 除外（確定№12）")
put("予約未処理数", reserved_pending)
put("フリー在庫数", free_stock, "現在庫(予約0)+入荷残-予約未処理（確定№3）")
put("フリー在庫日数", free_days, "分母=30日中央値に統一（仕様の平均記載は要確認）")
put("前回発注日", NA, "MMS発注書一覧/sitateru 未取得")
put("前回原価", NA, "予約管理表/sitateru 未取得")
put("画像", NA, "カラーコード(データ連携管理)未取得")
put("確定発注数", "", "手入力欄")
r += 1
put("▼直近7日 販売数", sales7)
put("▼直近7日 日販平均", velo7, "販売数÷7（7日未満は経過日数）")
put("▼直近7日 現在庫日数", sd7)
put("▼直近7日 完売想定日", soldout(sd7))
put("▼直近30日 販売数", sales30)
put("▼直近30日 日販中央値", median30, "在庫あり受注0=0/在庫無し受注0=除外")
put("▼直近30日 現在庫日数", sd30)
put("▼直近30日 完売想定日", soldout(sd30))
r += 1
section("入荷山（予約管理表）")
for i in range(3):
    if i < len(arrivals):
        put(f"入荷日{i+1}", asdate(arrivals[i].d)); put(f"　入荷数{i+1}", arrivals[i].q)
    else:
        put(f"入荷日{i+1}", ""); put(f"　入荷数{i+1}", "")
r += 1

section("FKU枚数構成（品番&カラー）")
for col, pct in sorted(fku_comp.items(), key=lambda x: -x[1]):
    put(f"　{col}", f"{pct}%", f"{fku[col]}枚")
r += 1

section("② 内訳")
put("合計販売数", tot_qty)
put("平均売価（税抜）", avg_price)
put("粗利率(%)", margin_pc)
put("値引率(%)", discount_pc)
put("予約販売数割合(%)", yoyaku_rate)
put("入荷数量（期間）", NA, "MMS発注書一覧 未取得（incoming_stockは現時点値）")
put("UU", UU, "sales_daily由来。CVRと1日ずれ")
put("CVR(%)", NA, "受注点数(ダッシュボード値)未配線")
put("お気に率(%)", NA, "お気に入り登録者数(ダッシュボード値)未配線")
put("CP対象枚数比(%)", NA, "クーポン実施日判定 要確認")
r += 1
section("時系列（注文数） yyyy")
for x in piv_y: put(f"　{int(x.k)}", x.q)
section("時系列（注文数） yyyy/mm")
for x in piv_m: put(f"　{x.k}", x.q)
section("時系列（注文数） yyyy/mm/dd")
for x in piv_d: put(f"　{x.k}", x.q)

ws.column_dimensions["A"].width = 26
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 46

outdir = r"C:\Users\Administrator\Downloads\system\exports"
os.makedirs(outdir, exist_ok=True)
outpath = os.path.join(outdir, f"発注管理表_{PC}_{START}_{END}.xlsx")
wb.save(outpath)
print(f"OK 出力: {outpath}")
print(f"   品番={PC} 期間={START}〜{END} 作成日={ASOF}")
print(f"   合計販売数={tot_qty} 平均売価={avg_price} 値引率={discount_pc}% 粗利率={margin_pc}%")
print(f"   現在庫={cur_stock} お気に入り={favorites} 予約未処理={reserved_pending} フリー在庫={free_stock}")
print(f"   7日販売={sales7} 日販平均={velo7} / 30日販売={sales30} 中央値={median30}")
print(f"   最終入荷日={last_arrival} レビュー件数={review_cnt} 点数={review_avg} UU={UU}")
