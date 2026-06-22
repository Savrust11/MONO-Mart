import sys, io, csv
from pathlib import Path
sys.path.insert(0, r"C:\Users\Administrator\Downloads\system\pipeline")
KEY = Path(r"C:\Users\Administrator\Downloads\system\pipeline\sheets-sa-key.json")
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from openpyxl import load_workbook

creds = Credentials.from_service_account_file(str(KEY), scopes=["https://www.googleapis.com/auth/drive.readonly"])
drive = build("drive","v3",credentials=creds,cache_discovery=False)
def find(fid,name):
    return drive.files().list(q=f"'{fid}' in parents and trashed=false and name='{name}'",
        fields="files(id,name)", supportsAllDrives=True, includeItemsFromAllDrives=True).execute().get("files",[])
def dl(fid):
    buf=io.BytesIO(); d=MediaIoBaseDownload(buf, drive.files().get_media(fileId=fid,supportsAllDrives=True)); done=False
    while not done: _,done=d.next_chunk()
    return buf.getvalue()

mf=find("1eRZoby0Bo8J4c6anWoe0PSGpDaGjePaR","202407")
of=find(mf[0]["id"],"2024_07_01.xlsx")
wb=load_workbook(io.BytesIO(dl(of[0]["id"])), read_only=True, data_only=True); ws=wb.active
hdr=[str(c).strip() if c is not None else "" for c in next(ws.iter_rows(min_row=1,max_row=1,values_only=True))]
# index of 注文数, 合計金額（税抜）, 注文日
for col in ("注文数","合計金額（税抜）","注文日","発送日"):
    idx = hdr.index(col) if col in hdr else -1
    print(f"col '{col}' at index {idx}")
oi = hdr.index("注文数")
print("\nFirst 6 data rows — 注文数 raw values:")
for i,row in enumerate(ws.iter_rows(min_row=2,max_row=7,values_only=True),start=2):
    print(f"  row{i}: 注文数={row[oi]!r}  合計={row[hdr.index('合計金額（税抜）')]!r}")
