import sys, io, csv
from pathlib import Path
sys.path.insert(0, r"C:\Users\Administrator\Downloads\system\pipeline")
KEY = Path(r"C:\Users\Administrator\Downloads\system\pipeline\sheets-sa-key.json")
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from openpyxl import load_workbook
from extractors.zozo_csv_extractor import ZOZOCsvExtractor

creds = Credentials.from_service_account_file(str(KEY), scopes=["https://www.googleapis.com/auth/drive.readonly"])
drive = build("drive","v3",credentials=creds,cache_discovery=False)
ext = ZOZOCsvExtractor()

def find(fid,name):
    return drive.files().list(q=f"'{fid}' in parents and trashed=false and name='{name}'",
        fields="files(id,name)", supportsAllDrives=True, includeItemsFromAllDrives=True).execute().get("files",[])
def dl(fid):
    buf=io.BytesIO(); d=MediaIoBaseDownload(buf, drive.files().get_media(fileId=fid,supportsAllDrives=True)); done=False
    while not done: _,done=d.next_chunk()
    return buf.getvalue()
def xlsx_to_csv_bytes(data):
    wb=load_workbook(io.BytesIO(data), read_only=True, data_only=True); ws=wb.active
    out=io.StringIO(); w=csv.writer(out)
    for row in ws.iter_rows(values_only=True):
        w.writerow(["" if c is None else c for c in row])
    return out.getvalue().encode("utf-8-sig")

def check(recs, label):
    print(f"\n=== {label}: {len(recs)} parsed records ===")
    if not recs: print("  NO RECORDS!"); return
    r=recs[0]
    for k in ("sale_date","sku_code","product_code","parent_category","sales_quantity","sales_amount","shipped_date","source_file"):
        print(f"  {k}: {r.get(k)}")
    qty=sum(x['sales_quantity'] for x in recs); amt=sum(x['sales_amount'] for x in recs)
    nd=len(set(x['sale_date'] for x in recs))
    print(f"  totals: qty={qty} amount={amt:.0f} distinct_dates={nd}")

# ORDERS (daily xlsx)
mf=find("1eRZoby0Bo8J4c6anWoe0PSGpDaGjePaR","202407")
of=find(mf[0]["id"],"2024_07_01.xlsx")
check(ext.parse_orders(xlsx_to_csv_bytes(dl(of[0]["id"])), "2024-07-01", is_shipped=False), "ORDERS 2024_07_01")

# SHIPPED (monthly xlsx)
sx=find("1kSuIgbkOtpfOu474A-rFRWKcLN0wKjJK","20240701(出荷).xlsx")
check(ext.parse_orders(xlsx_to_csv_bytes(dl(sx[0]["id"])), "2024-07-01", is_shipped=True), "SHIPPED 20240701 xlsx")

# SHIPPED (monthly csv)
scv=find("1kSuIgbkOtpfOu474A-rFRWKcLN0wKjJK","20240801(出荷).csv")
check(ext.parse_orders(dl(scv[0]["id"]), "2024-08-01", is_shipped=True), "SHIPPED 20240801 csv")
