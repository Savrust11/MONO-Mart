import sys, io, csv
from pathlib import Path
sys.path.insert(0, r"C:\Users\Administrator\Downloads\system\pipeline")
KEY = Path(r"C:\Users\Administrator\Downloads\system\pipeline\sheets-sa-key.json")
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from openpyxl import load_workbook

SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]
creds = Credentials.from_service_account_file(str(KEY), scopes=SCOPES)
drive = build("drive", "v3", credentials=creds, cache_discovery=False)

def find(fid, name):
    res = drive.files().list(q=f"'{fid}' in parents and trashed=false and name='{name}'",
        fields="files(id,name,mimeType)", supportsAllDrives=True,
        includeItemsFromAllDrives=True).execute()
    return res.get("files", [])

def dl(fid):
    buf = io.BytesIO()
    req = drive.files().get_media(fileId=fid, supportsAllDrives=True)
    d = MediaIoBaseDownload(buf, req)
    done = False
    while not done:
        _, done = d.next_chunk()
    return buf.getvalue()

def show_xlsx(data, label):
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    print(f"\n--- {label} (xlsx, sheet={ws.title}) ---")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        cells = [str(c)[:16] if c is not None else "" for c in row[:12]]
        print(f"  row{i}: {cells}")

def show_csv(data, label):
    for enc in ("utf-8-sig","cp932","shift_jis"):
        try:
            text = data.decode(enc); break
        except UnicodeDecodeError: continue
    rows = list(csv.reader(io.StringIO(text)))
    print(f"\n--- {label} (csv, enc={enc}, {len(rows)} rows) ---")
    for i, r in enumerate(rows[:5], start=1):
        print(f"  row{i}: {[c[:16] for c in r[:12]]}")

# orders: dive into 202407 then 2024_07_01.xlsx
mf = find("1eRZoby0Bo8J4c6anWoe0PSGpDaGjePaR", "202407")
of = find(mf[0]["id"], "2024_07_01.xlsx")
show_xlsx(dl(of[0]["id"]), "ORDERS 2024_07_01.xlsx")

# shipped: the xlsx one and a csv one
sx = find("1kSuIgbkOtpfOu474A-rFRWKcLN0wKjJK", "20240701(出荷).xlsx")
show_xlsx(dl(sx[0]["id"]), "SHIPPED 20240701(出荷).xlsx")
sc = find("1kSuIgbkOtpfOu474A-rFRWKcLN0wKjJK", "20240801(出荷).csv")
show_csv(dl(sc[0]["id"]), "SHIPPED 20240801(出荷).csv")
