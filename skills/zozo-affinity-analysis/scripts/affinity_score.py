"""
ZOZO買い回り相性分析 Step 3: 相性スコア算出・Excel出力
merged_other.csv を入力として相性スコアを算出し、書式付きExcelを出力する。

Usage:
  python affinity_score.py \
    --merged_other <merged_other.csv> \
    --output <相性分析レポート.xlsx> \
    [--min_buyback 30] \
    [--min_buyers 100] \
    [--top_n 20]
"""
import argparse
import pandas as pd
import numpy as np
import warnings
warnings.filterwarnings('ignore')
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 自社ブランド一覧（シート生成対象）
OWN_BRANDS = [
    'MONO-MART', 'EMMA CLOTHES', 'WYM LIDNM', 'THE CRAFT CREW',
    'Alfred Alex', 'Anchor Smith', 'ADRER', 'CLEL', 'LOOSE', 'cussil',
    'GRANCY', 'SERACE', 'LUENNA', 'RUUBON', 'forksy.', "MONO-MART LADY'S",
    'BONLECILL', 'Heart Tattoo', 'ELUNIS', 'Elishe', 'Parts Lab.', 'Aunely'
]


def calc_affinity(df, min_buyback, min_buyers):
    """相性スコアを算出してフィルタ済みDataFrameを返す。"""
    d = df.copy()
    d = d[d['買い回り件数'] >= min_buyback]
    d = d[d['購入者数'] >= min_buyers]
    d['相性スコア(%)'] = (d['買い回り件数'] / d['購入者数'] * 100).round(4)
    return d


def format_brand_df(bdf):
    """ブランド別DataFrameを出力用に整形する。"""
    out = bdf[['ショップ名', '買い回り件数', '購入者数', '相性スコア(%)',
               '受注売上', '1枚単価', '客単価',
               '購入者男性比率', '購入者女性比率', '購入者平均年齢', '売上昨対比']].copy()
    out.columns = ['ショップ名', '買い回り件数', 'ショップ購入者数', '相性スコア(%)',
                   'ショップ受注売上(円)', 'ショップ1枚単価(円)', 'ショップ客単価(円)',
                   '男性比率', '女性比率', '平均年齢', '売上昨対比']
    out['男性比率'] = (out['男性比率'] * 100).round(1)
    out['女性比率'] = (out['女性比率'] * 100).round(1)
    out['売上昨対比'] = (out['売上昨対比'] * 100).round(1)
    return out


def apply_excel_style(wb):
    """Excelブックに書式を適用する。"""
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    alt_fill = PatternFill(start_color='EBF3FB', end_color='EBF3FB', fill_type='solid')
    border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            fill = alt_fill if i % 2 == 0 else PatternFill()
            for cell in row:
                cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if cell.column == 1:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    val = str(cell.value) if cell.value else ''
                    jlen = sum(2 if ord(c) > 127 else 1 for c in val)
                    max_len = max(max_len, jlen)
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
        ws.row_dimensions[1].height = 30
        for i in range(2, ws.max_row + 1):
            ws.row_dimensions[i].height = 18
        ws.freeze_panes = 'B2'
    return wb


def main():
    parser = argparse.ArgumentParser(description='ZOZO買い回り相性スコア算出・Excel出力')
    parser.add_argument('--merged_other', required=True, help='merged_other.csv パス')
    parser.add_argument('--output', required=True, help='出力Excelパス')
    parser.add_argument('--min_buyback', type=int, default=30, help='最低買い回り件数（デフォルト30）')
    parser.add_argument('--min_buyers', type=int, default=100, help='最低ショップ購入者数（デフォルト100）')
    parser.add_argument('--top_n', type=int, default=20, help='ブランド別上位N件（デフォルト20）')
    args = parser.parse_args()

    df = pd.read_csv(args.merged_other, encoding='utf-8-sig')
    scored = calc_affinity(df, args.min_buyback, args.min_buyers)
    print(f"フィルタ後レコード数: {len(scored)}")

    # ブランド別集計
    results = {}
    for brand in OWN_BRANDS:
        bdf = scored[scored['ブランド'] == brand].sort_values('相性スコア(%)', ascending=False).head(args.top_n)
        if len(bdf) == 0:
            continue
        results[brand] = format_brand_df(bdf)

    # 全ブランド横断
    cross = scored.groupby('ショップ名').agg(
        対象ブランド数=('ブランド', 'nunique'),
        平均相性スコア=('相性スコア(%)', 'mean'),
        合計買い回り件数=('買い回り件数', 'sum'),
        ショップ購入者数=('購入者数', 'first'),
        ショップ受注売上=('受注売上', 'first'),
        ショップ1枚単価=('1枚単価', 'first'),
        ショップ客単価=('客単価', 'first'),
        男性比率=('購入者男性比率', 'first'),
        女性比率=('購入者女性比率', 'first'),
        平均年齢=('購入者平均年齢', 'first'),
        売上昨対比=('売上昨対比', 'first'),
        相性ブランド一覧=('ブランド', lambda x: ' / '.join(sorted(x.unique())))
    ).reset_index()
    cross['平均相性スコア(%)'] = cross['平均相性スコア'].round(4)
    cross['男性比率'] = (cross['男性比率'] * 100).round(1)
    cross['女性比率'] = (cross['女性比率'] * 100).round(1)
    cross['売上昨対比'] = (cross['売上昨対比'] * 100).round(1)
    cross = cross.sort_values('平均相性スコア(%)', ascending=False)

    multi = cross[cross['対象ブランド数'] >= 3].head(30)

    # Excel出力
    with pd.ExcelWriter(args.output, engine='openpyxl') as writer:
        # 複数ブランド相性良TOP30
        multi_out = multi[['ショップ名', '対象ブランド数', '平均相性スコア(%)', '合計買い回り件数',
                            'ショップ購入者数', 'ショップ受注売上', 'ショップ1枚単価', 'ショップ客単価',
                            '男性比率', '女性比率', '平均年齢', '売上昨対比', '相性ブランド一覧']].copy()
        multi_out.columns = ['ショップ名', '相性ブランド数', '平均相性スコア(%)', '合計買い回り件数',
                              'ショップ購入者数', 'ショップ受注売上(円)', 'ショップ1枚単価(円)', 'ショップ客単価(円)',
                              '男性比率(%)', '女性比率(%)', '平均年齢', '売上昨対比(%)', '相性ブランド一覧']
        multi_out.to_excel(writer, sheet_name='★複数ブランド相性良TOP30', index=False)

        # 全ブランド横断TOP50
        cross_out = cross[['ショップ名', '対象ブランド数', '平均相性スコア(%)', '合計買い回り件数',
                            'ショップ購入者数', 'ショップ受注売上', 'ショップ1枚単価', 'ショップ客単価',
                            '男性比率', '女性比率', '平均年齢', '売上昨対比']].head(50).copy()
        cross_out.columns = ['ショップ名', '相性ブランド数', '平均相性スコア(%)', '合計買い回り件数',
                              'ショップ購入者数', 'ショップ受注売上(円)', 'ショップ1枚単価(円)', 'ショップ客単価(円)',
                              '男性比率(%)', '女性比率(%)', '平均年齢', '売上昨対比(%)']
        cross_out.to_excel(writer, sheet_name='全ブランド横断TOP50', index=False)

        # ブランド別シート
        for brand, bdf in results.items():
            bdf.to_excel(writer, sheet_name=brand[:31], index=False)

    # 書式適用
    wb = load_workbook(args.output)
    wb = apply_excel_style(wb)
    wb.save(args.output)

    print(f"\nExcel出力完了: {args.output}")
    print(f"シート数: {len(wb.sheetnames)}")

    # コンソール：主要ブランド上位5
    key_brands = ['MONO-MART', 'CLEL', 'ADRER', 'EMMA CLOTHES', 'WYM LIDNM']
    for brand in key_brands:
        if brand not in results:
            continue
        print(f"\n【{brand}】相性スコア上位5")
        for i, row in results[brand].head(5).iterrows():
            print(f"  {row['相性スコア(%)']:.2f}% | {row['ショップ名']} | 買い回り{int(row['買い回り件数'])}件")


if __name__ == '__main__':
    main()
