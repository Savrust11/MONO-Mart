#!/usr/bin/env python3
"""
ZOZOAD レポート更新スクリプト
CSVデータを整形してGoogle Drive上のxlsxに新しいシートとして追加する。

Usage:
    python3 update_zozoad_sheet.py <csv_path> <date_mmdd>

Example:
    python3 update_zozoad_sheet.py /home/ubuntu/Downloads/report.csv 0411
"""

import sys
import os
import subprocess
import csv
import tempfile
from openpyxl import load_workbook
from openpyxl.styles import Font

# 定数
SPREADSHEET_FILE_ID = "1-X0vkCaTgycj5qIXArofLKBqDbqob05Z"
SPREADSHEET_URL = "https://docs.google.com/spreadsheets/d/1-X0vkCaTgycj5qIXArofLKBqDbqob05Z/edit"

# ショップの表示順序
SHOP_ORDER = ["MONO-MART", "EMMA CLOTHES", "ADRER", "Anchor Smith", "BONLECILL"]

# ヘッダー列名（18列）
HEADERS = [
    "ショップID", "ショップ名", "親カテゴリ", "子カテゴリ",
    "ブランド品番", "商品コード", "商品名", "親商品タイプ",
    "子商品タイプ", "アップロード日", "imp", "click",
    "コスト", "経由売上件数", "経由売上金額（税抜）", "CTR", "CPC", "ROAS"
]

# click列のインデックス（0始まり）
CLICK_COL_INDEX = 11


def download_xlsx(output_path):
    """Google DriveからxlsxをDL"""
    cmd = [
        "gws", "drive", "files", "get",
        "--params", f'{{"fileId": "{SPREADSHEET_FILE_ID}", "alt": "media"}}',
        "--output", output_path
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        print(f"Error downloading xlsx: {result.stderr}")
        sys.exit(1)
    print(f"Downloaded xlsx to {output_path}")


def upload_xlsx(input_path):
    """xlsxをGoogle Driveに再アップロード（上書き更新）"""
    cmd = [
        "gws", "drive", "files", "update",
        "--params", f'{{"fileId": "{SPREADSHEET_FILE_ID}"}}',
        "--upload", input_path,
        "--upload-content-type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        print(f"Error uploading xlsx: {result.stderr}")
        sys.exit(1)
    print(f"Uploaded xlsx from {input_path}")
    print(f"Spreadsheet URL: {SPREADSHEET_URL}")


def read_csv(csv_path):
    """CSVを読み込んでショップごとにグループ化する"""
    # エンコーディングを自動判定（cp932 → utf-8 の順で試行）
    for encoding in ["cp932", "utf-8", "utf-8-sig"]:
        try:
            with open(csv_path, "r", encoding=encoding) as f:
                reader = csv.reader(f)
                rows = list(reader)
            print(f"CSV read with encoding: {encoding}, rows: {len(rows)}")
            break
        except (UnicodeDecodeError, UnicodeError):
            continue
    else:
        print("Error: Could not decode CSV with any known encoding")
        sys.exit(1)

    if len(rows) < 2:
        print("Error: CSV has no data rows")
        sys.exit(1)

    # ヘッダー行をスキップ（1行目）
    csv_header = rows[0]
    data_rows = rows[1:]

    # ショップ名列のインデックスを特定
    shop_col = None
    for i, h in enumerate(csv_header):
        if "ショップ名" in h:
            shop_col = i
            break
    if shop_col is None:
        # デフォルトは列2（0始まりで1）
        shop_col = 1

    # click列のインデックスを特定
    click_col = None
    for i, h in enumerate(csv_header):
        if h.strip() == "click":
            click_col = i
            break
    if click_col is None:
        click_col = CLICK_COL_INDEX

    # ショップごとにグループ化
    shop_data = {}
    for row in data_rows:
        if len(row) < len(csv_header):
            # 短い行はスキップ
            continue
        shop_name = row[shop_col].strip() if shop_col < len(row) else ""
        if not shop_name:
            continue
        if shop_name not in shop_data:
            shop_data[shop_name] = []
        shop_data[shop_name].append(row)

    # 各ショップ内をclick降順でソート
    for shop_name in shop_data:
        shop_data[shop_name].sort(
            key=lambda r: _to_number(r[click_col] if click_col < len(r) else "0"),
            reverse=True
        )

    print(f"Shops found: {list(shop_data.keys())}")
    for shop, rows in shop_data.items():
        print(f"  {shop}: {len(rows)} rows")

    return shop_data


def _to_number(val):
    """文字列を数値に変換（変換できない場合は0）"""
    try:
        return float(str(val).replace(",", "").strip())
    except (ValueError, TypeError):
        return 0


def _convert_value(val, col_index):
    """CSV値をExcelに適した型に変換する"""
    if val is None or val == "":
        return None

    val = str(val).strip()

    # 数値列（imp, click, コスト, 経由売上件数, 経由売上金額, CTR, CPC, ROAS, ショップID, 商品コード）
    numeric_cols = {0, 5, 10, 11, 12, 13, 14, 15, 16, 17}
    if col_index in numeric_cols:
        try:
            cleaned = val.replace(",", "")
            if "." in cleaned:
                return float(cleaned)
            else:
                return int(cleaned)
        except (ValueError, TypeError):
            return val

    return val


def add_sheet_to_xlsx(xlsx_path, sheet_name, shop_data):
    """xlsxに新しいシートを追加する"""
    wb = load_workbook(xlsx_path)

    # 同名シートが既に存在する場合は削除
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(title=sheet_name)

    bold_font = Font(bold=True)
    current_row = 1

    # ショップ順序に従って書き込み
    first_shop = True
    for shop_name in SHOP_ORDER:
        if shop_name not in shop_data:
            continue

        if not first_shop:
            # ショップ間に2行の空行
            current_row += 2

        # ヘッダー行（太字）
        for col_idx, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = bold_font
        current_row += 1

        # データ行
        for row_data in shop_data[shop_name]:
            for col_idx, val in enumerate(row_data[:18], 1):
                converted = _convert_value(val, col_idx - 1)
                ws.cell(row=current_row, column=col_idx, value=converted)
            current_row += 1

        first_shop = False

    # SHOP_ORDERに含まれないショップも追加
    for shop_name in shop_data:
        if shop_name in SHOP_ORDER:
            continue

        if not first_shop:
            current_row += 2

        for col_idx, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = bold_font
        current_row += 1

        for row_data in shop_data[shop_name]:
            for col_idx, val in enumerate(row_data[:18], 1):
                converted = _convert_value(val, col_idx - 1)
                ws.cell(row=current_row, column=col_idx, value=converted)
            current_row += 1

        first_shop = False

    wb.save(xlsx_path)
    print(f"Added sheet '{sheet_name}' with {current_row - 1} rows")


def main():
    if len(sys.argv) < 3:
        print("Usage: python3 update_zozoad_sheet.py <csv_path> <date_mmdd>")
        print("Example: python3 update_zozoad_sheet.py /home/ubuntu/Downloads/report.csv 0411")
        sys.exit(1)

    csv_path = sys.argv[1]
    date_mmdd = sys.argv[2]

    if not os.path.exists(csv_path):
        print(f"Error: CSV file not found: {csv_path}")
        sys.exit(1)

    # Step 1: Google DriveからxlsxをDL
    xlsx_path = "/home/ubuntu/zozoad_data_work.xlsx"
    print("=== Step 1: Downloading xlsx from Google Drive ===")
    download_xlsx(xlsx_path)

    # Step 2: CSVを読み込み・整形
    print(f"\n=== Step 2: Reading CSV: {csv_path} ===")
    shop_data = read_csv(csv_path)

    # Step 3: xlsxに新しいシートを追加
    print(f"\n=== Step 3: Adding sheet '{date_mmdd}' to xlsx ===")
    add_sheet_to_xlsx(xlsx_path, date_mmdd, shop_data)

    # Step 4: Google Driveに再アップロード
    print(f"\n=== Step 4: Uploading xlsx to Google Drive ===")
    upload_xlsx(xlsx_path)

    # クリーンアップ
    os.remove(xlsx_path)
    print(f"\n=== Done! Sheet '{date_mmdd}' added successfully ===")
    print(f"Spreadsheet URL: {SPREADSHEET_URL}")


if __name__ == "__main__":
    main()
