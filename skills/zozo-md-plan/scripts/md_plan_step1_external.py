#!/usr/bin/env python3
"""
MD計画 Step1: 外部環境分析
買い回りデータ（3ヶ月分）× ショップ指標データ（3ヶ月分）

出力:
- ブランド別 × カテゴリ（商品親タイプ）× 価格帯の買い回り集計
- ショップ別相性スコア（母数補正済み）
- カテゴリ別相性スコア
"""

import pandas as pd
import numpy as np
import json
import os
import gc
from collections import defaultdict

# ===== 設定 =====
KAIMAWARI_FILES = [
    '/home/ubuntu/買い回りデータ2026年1月.xlsx',
    '/home/ubuntu/買い回りデータ2026年2月.xlsx',
    '/home/ubuntu/買い回りデータ2026年3月.xlsx',
]
SHOP_FILES = [
    '/home/ubuntu/shop_data/2026年1月集計.xls',
    '/home/ubuntu/shop_data/2026年2月集計.xls',
    '/home/ubuntu/shop_data/2026年3月集計.xls',
]

# 分析対象ブランド（主要5ブランド）
TARGET_BRANDS = ['MONO-MART', 'EMMA CLOTHES', 'CLEL', 'ADRER', 'WYM LIDNM']

# 自社ブランドリスト（除外用）
OWN_BRANDS = [
    'MONO-MART', 'EMMA CLOTHES', 'WYM LIDNM', 'THE CRAFT CREW',
    'Alfred Alex', 'Anchor Smith', 'ADRER', 'CLEL', 'LOOSE', 'cussil',
    'GRANCY', 'SERACE', 'LUENNA', 'RUUBON', 'forksy.', "MONO-MART LADY'S",
    'BONLECILL', 'Heart Tattoo', 'ELUNIS', 'Elishe', 'Parts Lab.', 'Aunely',
    'THE CRAFT CREW PRODUCTS',
]

# 価格帯定義
PRICE_BANDS = [
    (0, 1999, '〜1,999円'),
    (2000, 3999, '2,000〜3,999円'),
    (4000, 5999, '4,000〜5,999円'),
    (6000, 7999, '6,000〜7,999円'),
    (8000, 11999, '8,000〜11,999円'),
    (12000, float('inf'), '12,000円〜'),
]

def get_price_band(price):
    """価格帯を返す"""
    try:
        p = float(price)
    except (ValueError, TypeError):
        return '不明'
    for low, high, label in PRICE_BANDS:
        if low <= p <= high:
            return label
    return '不明'

def load_shop_buyers():
    """ショップ指標データから月別購入者数を読み込み、月平均を算出"""
    shop_monthly = defaultdict(list)
    
    for fpath in SHOP_FILES:
        print(f'  Loading shop data: {os.path.basename(fpath)}')
        df = pd.read_excel(fpath, header=1)
        df = df[['ショップ名', '購入者数']].dropna(subset=['ショップ名', '購入者数'])
        df['購入者数'] = pd.to_numeric(df['購入者数'], errors='coerce')
        df = df.dropna(subset=['購入者数'])
        
        for _, row in df.iterrows():
            shop_monthly[row['ショップ名']].append(row['購入者数'])
    
    # 月平均購入者数を算出
    shop_avg_buyers = {}
    for shop, buyers_list in shop_monthly.items():
        shop_avg_buyers[shop] = np.mean(buyers_list)
    
    print(f'  Total shops with buyer data: {len(shop_avg_buyers)}')
    return shop_avg_buyers

def process_kaimawari_brand(brand, kaimawari_files):
    """1ブランドの買い回りデータを全月分読み込み・集計（メモリ効率化）"""
    
    # 集計用辞書
    shop_agg = defaultdict(lambda: {'件数': 0, '点数': 0, '金額': 0})
    cat_agg = defaultdict(lambda: {'件数': 0, '点数': 0, '金額': 0})
    cat_price_agg = defaultdict(lambda: {'件数': 0, '点数': 0, '金額': 0})
    
    for fpath in kaimawari_files:
        print(f'  Processing {brand} from {os.path.basename(fpath)}...')
        try:
            # openpyxlでシートを読み込み（メモリ効率のためチャンク処理）
            import openpyxl
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            
            if brand not in wb.sheetnames:
                print(f'    Sheet "{brand}" not found, skipping')
                wb.close()
                continue
            
            ws = wb[brand]
            rows = ws.iter_rows(values_only=True)
            header = next(rows)  # Skip header
            
            # Column indices:
            # [3] ショップ名, [4] 親カテゴリ(=brand), [5] 子カテゴリ(=brand)
            # [6] 商品親タイプ(=category), [7] 商品子タイプ(=sub-category)
            # [12] 上代, [13] 販売価格, [16] 売上点数, [17] 売上金額
            
            row_count = 0
            for row in rows:
                if row is None or len(row) < 18:
                    continue
                
                shop_name = row[3]
                category = row[6]  # 商品親タイプ = actual category
                price = row[13]    # 販売価格
                points = row[16]   # 売上点数
                amount = row[17]   # 売上金額
                
                if shop_name is None or category is None:
                    continue
                
                # 自社ブランドのショップを除外
                if any(own in str(shop_name) for own in OWN_BRANDS):
                    continue
                
                try:
                    pts = int(points) if points else 0
                    amt = float(amount) if amount else 0
                except (ValueError, TypeError):
                    pts = 0
                    amt = 0
                
                price_band = get_price_band(price)
                
                # ショップ別集計
                shop_agg[str(shop_name)]['件数'] += 1
                shop_agg[str(shop_name)]['点数'] += pts
                shop_agg[str(shop_name)]['金額'] += amt
                
                # カテゴリ別集計
                cat_agg[str(category)]['件数'] += 1
                cat_agg[str(category)]['点数'] += pts
                cat_agg[str(category)]['金額'] += amt
                
                # カテゴリ×価格帯集計
                key = (str(category), price_band)
                cat_price_agg[key]['件数'] += 1
                cat_price_agg[key]['点数'] += pts
                cat_price_agg[key]['金額'] += amt
                
                row_count += 1
            
            wb.close()
            print(f'    Processed {row_count:,} rows (other-brand)')
            
        except Exception as e:
            print(f'    Error: {e}')
    
    gc.collect()
    return shop_agg, cat_agg, cat_price_agg

def main():
    print('=== MD計画 Step1: 外部環境分析 ===')
    
    # 1. ショップ購入者数を読み込み
    print('\n[1] Loading shop buyer data...')
    shop_avg_buyers = load_shop_buyers()
    
    # 2. ブランド別に買い回りデータを処理
    result = {
        'affinity': {},           # ショップ別相性スコア
        'category_kaimawari': {}, # カテゴリ別買い回り
        'category_price_matrix': {},  # カテゴリ×価格帯
    }
    
    for brand in TARGET_BRANDS:
        print(f'\n[2] Processing brand: {brand}')
        shop_agg, cat_agg, cat_price_agg = process_kaimawari_brand(brand, KAIMAWARI_FILES)
        
        # ショップ別相性スコア算出
        shop_scores = []
        for shop, data in shop_agg.items():
            avg_buyers = shop_avg_buyers.get(shop, 0)
            if avg_buyers >= 100 and data['件数'] >= 30:
                score = (data['件数'] / avg_buyers) * 100
                shop_scores.append({
                    'ショップ名': shop,
                    '買い回り件数': data['件数'],
                    '買い回り点数': data['点数'],
                    '買い回り金額': data['金額'],
                    '月平均購入者数': round(avg_buyers, 1),
                    '相性スコア': round(score, 2),
                })
        
        # スコア降順でTOP30
        shop_scores.sort(key=lambda x: x['相性スコア'], reverse=True)
        result['affinity'][brand] = shop_scores[:30]
        
        # カテゴリ別集計
        cat_list = []
        for cat, data in cat_agg.items():
            if data['件数'] >= 10:
                avg_price = data['金額'] / data['点数'] if data['点数'] > 0 else 0
                cat_list.append({
                    'カテゴリ': cat,
                    '件数': data['件数'],
                    '点数': data['点数'],
                    '金額': data['金額'],
                    '平均単価': round(avg_price),
                    'ショップ数': len([s for s, d in shop_agg.items() if d['件数'] >= 1]),
                })
        cat_list.sort(key=lambda x: x['件数'], reverse=True)
        result['category_kaimawari'][brand] = cat_list[:30]
        
        # カテゴリ×価格帯マトリクス
        cp_list = []
        for (cat, pb), data in cat_price_agg.items():
            if data['件数'] >= 5:
                cp_list.append({
                    'カテゴリ': cat,
                    '価格帯': pb,
                    '件数': data['件数'],
                    '点数': data['点数'],
                    '金額': data['金額'],
                })
        cp_list.sort(key=lambda x: x['件数'], reverse=True)
        result['category_price_matrix'][brand] = cp_list
        
        print(f'  → {brand}: {len(shop_scores)} shops scored, {len(cat_list)} categories, {len(cp_list)} cat×price combos')
        gc.collect()
    
    # 3. 結果を保存
    output_path = '/home/ubuntu/md_plan_step1_result.json'
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    
    print(f'\n=== Step1 完了: {output_path} ===')
    
    # サマリー表示
    for brand in TARGET_BRANDS:
        aff = result['affinity'].get(brand, [])
        cats = result['category_kaimawari'].get(brand, [])
        print(f'\n{brand}:')
        if aff:
            print(f'  TOP3 相性ショップ: {", ".join(a["ショップ名"] for a in aff[:3])}')
        if cats:
            print(f'  TOP5 カテゴリ: {", ".join(c["カテゴリ"] for c in cats[:5])}')

if __name__ == '__main__':
    main()
