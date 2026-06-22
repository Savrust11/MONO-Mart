#!/usr/bin/env python3
"""
ZOZO 52週カテゴリ別売上係数を算出するスクリプト。

Usage:
    python build_coefficients.py <zozo_52week_xlsx> <output_json>

Args:
    zozo_52week_xlsx: 【52週】カテゴリ推移.xlsx のパス
    output_json: 出力先JSONファイルパス

Output:
    カテゴリ(子)ごとの52週係数JSON。各週の値は年間合計に対する構成比。
"""
import sys
import json
from collections import defaultdict
import openpyxl

def build_coefficients(xlsx_path, output_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # 「ZOZO全体」シートを探す
    sheet_name = None
    for name in wb.sheetnames:
        if 'ZOZO' in name and '全体' in name:
            sheet_name = name
            break
    if not sheet_name:
        # フォールバック: PVTシートを探す
        for name in wb.sheetnames:
            if 'PVT' in name.upper() or 'サマリ' in name:
                sheet_name = name
                break
    if not sheet_name:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]

    # ヘッダー行を特定
    # 既知フォーマット: Row1がヘッダー、列名で自動検出
    header_row = 1
    cat_col = None
    week_col = None
    sales_col = None
    filter_col = None  # 「ZOZO全体」フィルタ用

    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        for col_idx, val in enumerate(row):
            if val is None:
                continue
            val_str = str(val)
            # カテゴリ(子) or 商品タイプ(子)
            if ('カテゴリ' in val_str or '商品タイプ' in val_str) and '子' in val_str:
                cat_col = col_idx
            elif val_str in ('週', '週数'):
                week_col = col_idx
            elif val_str in ('販売数', '売上数'):
                sales_col = col_idx
            elif '全体' in val_str and '社内' in val_str:
                filter_col = col_idx

    if cat_col is None or week_col is None or sales_col is None:
        print("Error: 必要な列が見つかりません。")
        print(f"シート: {sheet_name}, cat_col={cat_col}, week_col={week_col}, sales_col={sales_col}")
        wb.close()
        sys.exit(1)

    print(f"シート: {sheet_name}")
    print(f"ヘッダー行: {header_row}, カテゴリ列: {cat_col}, 週列: {week_col}, 販売数列: {sales_col}, フィルタ列: {filter_col}")

    # データ集計: カテゴリ×週 → 販売数合計
    cat_week_sales = defaultdict(lambda: defaultdict(int))
    cat_total = defaultdict(int)

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        cat = row[cat_col] if cat_col < len(row) else None
        week = row[week_col] if week_col is not None and week_col < len(row) else None
        sales = row[sales_col] if sales_col is not None and sales_col < len(row) else None

        if cat is None or week is None or sales is None:
            continue

        # 「ZOZO全体」行のみ集計（フィルタ列がある場合）
        if filter_col is not None:
            filter_val = row[filter_col] if filter_col < len(row) else None
            if filter_val and '全体' not in str(filter_val):
                continue
        try:
            week = int(week)
            sales = int(sales)
        except (ValueError, TypeError):
            continue

        if 1 <= week <= 52:
            cat_week_sales[str(cat)][week] += sales
            cat_total[str(cat)] += sales

    # 係数算出
    coefficients = {}
    for cat in sorted(cat_week_sales.keys()):
        total = cat_total[cat]
        if total <= 0:
            continue
        weekly = {}
        for w in range(1, 53):
            weekly[str(w)] = cat_week_sales[cat].get(w, 0) / total
        coefficients[cat] = {
            "total_sales": total,
            "weekly_coefficients": weekly
        }

    wb.close()

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(coefficients, f, ensure_ascii=False, indent=2)

    print(f"カテゴリ数: {len(coefficients)}")
    print(f"出力: {output_path}")

if __name__ == '__main__':
    if len(sys.argv) != 3:
        print(f"Usage: python {sys.argv[0]} <zozo_52week_xlsx> <output_json>")
        sys.exit(1)
    build_coefficients(sys.argv[1], sys.argv[2])
