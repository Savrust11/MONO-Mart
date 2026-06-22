#!/usr/bin/env python3
"""
在庫消化予測スクリプト。

品番別の在庫データと52週カテゴリ係数から、指定期日までの在庫消化を予測する。

Usage:
    python forecast_inventory.py <inventory_xlsx> <coefficients_json> <target_date> [--sheet SHEET] [--targets CODES] [--recent-days N] [--output CSV]

Args:
    inventory_xlsx:     品番別在庫・日別販売データのExcel（MONO-MART形式）
    coefficients_json:  build_coefficients.py で生成した52週係数JSON
    target_date:        予測期限（YYYY-MM-DD形式、例: 2026-05-31）

Options:
    --sheet SHEET       対象シート名（デフォルト: 自動検出）
    --targets CODES     対象品番をカンマ区切りで指定（省略時: 全品番）
    --recent-days N     日販計算に使う直近日数（デフォルト: 14）
    --output CSV        出力CSVパス（デフォルト: forecast_results.csv）

Excel想定フォーマット:
    Row 3: ヘッダー行
    A: ブランド, B: ブランド品番, C: 商品名, D: 元上代, E: 下代単価
    F: カテゴリ(親), G: カテゴリ(子), H: シーズン, I: 種類, J: テイスト
    K: 26SS入荷残, L: 26AW入荷残, M: 在庫数, N: 売れた合計
    O以降: 日別販売数（日付がヘッダーに入っている）
"""
import sys
import json
import csv
import argparse
from datetime import datetime, timedelta
from collections import defaultdict
import openpyxl


def get_week_number(dt):
    """年初からの週番号（1-52）を返す。1/1～1/7が第1週。"""
    day_of_year = dt.timetuple().tm_yday
    week = (day_of_year - 1) // 7 + 1
    return min(week, 52)


def detect_date_columns(ws, header_row=3):
    """ヘッダー行から日付列の開始・終了インデックスと日付リストを検出する。"""
    dates = []
    start_idx = None
    for row in ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True):
        for i, val in enumerate(row):
            if isinstance(val, datetime):
                if start_idx is None:
                    start_idx = i
                dates.append(val)
    end_idx = start_idx + len(dates) - 1 if start_idx is not None else None
    return start_idx, end_idx, dates


def load_items(ws, target_set=None, recent_days=14, header_row=3):
    """Excelシートから品番データを読み込む。"""
    start_idx, end_idx, dates = detect_date_columns(ws, header_row)
    if start_idx is None:
        print("Error: 日付列が検出できません。")
        sys.exit(1)

    # 実際にデータが入っている最終日を特定（0でない最後の日）
    # → ヘッダーの日付から「現在日」を推定
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    items = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        item_code = row[1]
        if item_code is None:
            continue
        item_code = str(item_code)
        if target_set and item_code not in target_set:
            continue

        stock = row[12] if row[12] else 0
        total_sold = row[13] if row[13] else 0

        # 日別販売データを取得
        daily_sales = []
        for i in range(start_idx, end_idx + 1):
            val = row[i] if i < len(row) and row[i] is not None else 0
            daily_sales.append(val)

        # 実データがある日数を特定（末尾の連続0を除外）
        actual_days = len(daily_sales)
        while actual_days > 0 and daily_sales[actual_days - 1] == 0:
            actual_days -= 1

        if actual_days == 0:
            avg_daily = 0.0
        else:
            # 直近N日の日販平均
            recent_start = max(0, actual_days - recent_days)
            recent_slice = daily_sales[recent_start:actual_days]
            avg_daily = sum(recent_slice) / len(recent_slice) if recent_slice else 0.0

            # 直近が0なら全期間平均にフォールバック
            if avg_daily <= 0:
                avg_daily = sum(daily_sales[:actual_days]) / actual_days

        # 最終データ日を推定
        if actual_days > 0 and dates:
            last_data_date = dates[actual_days - 1]
        else:
            last_data_date = today

        items.append({
            'brand': row[0],
            'item_code': item_code,
            'item_name': row[2],
            'price': row[3],
            'cat_parent': row[5],
            'cat_child': row[6],
            'season': row[7],
            'stock': stock,
            'total_sold': total_sold,
            'avg_daily': avg_daily,
            'last_data_date': last_data_date,
        })

    return items


def forecast(items, coefficients, current_date, target_date):
    """各品番の在庫消化予測を計算する。"""
    current_week = get_week_number(current_date)
    results = []

    for item in items:
        stock = item['stock']
        cat_child = item['cat_child']
        avg_daily = item['avg_daily']

        if cat_child and str(cat_child) in coefficients:
            coef_data = coefficients[str(cat_child)]['weekly_coefficients']
            current_coef = coef_data.get(str(current_week), 0)

            total_predicted = 0
            day = current_date
            while day <= target_date:
                wk = get_week_number(day)
                wk_coef = coef_data.get(str(wk), 0)
                ratio = wk_coef / current_coef if current_coef > 0 else 1.0
                predicted_daily = avg_daily * ratio

                days_count = 0
                temp = day
                while temp <= target_date and get_week_number(temp) == wk:
                    days_count += 1
                    temp += timedelta(days=1)

                total_predicted += predicted_daily * days_count
                day = temp
        else:
            days_remaining = (target_date - current_date).days + 1
            total_predicted = avg_daily * days_remaining

        remaining_stock = stock - total_predicted
        consumption_rate = (total_predicted / stock * 100) if stock > 0 else 0

        results.append({
            'brand': item['brand'],
            'item_code': item['item_code'],
            'item_name': item['item_name'],
            'cat_child': cat_child,
            'season': item['season'],
            'stock': stock,
            'avg_daily': round(avg_daily, 1),
            'predicted_sales': round(total_predicted),
            'remaining_stock': round(remaining_stock),
            'consumption_rate': round(consumption_rate, 1),
        })

    results.sort(key=lambda x: -x['remaining_stock'])
    return results


def output_results(results, output_path):
    """結果をCSVに出力し、サマリを表示する。"""
    with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['ブランド', '品番', '商品名', 'カテゴリ(子)', 'シーズン',
                         '現在庫', '直近日販', '予測販売数', '残在庫', '消化率(%)', '判定'])
        for r in results:
            judge = '余剰' if r['remaining_stock'] > 0 else '完売見込'
            writer.writerow([r['brand'], r['item_code'], r['item_name'],
                            r['cat_child'], r['season'], r['stock'],
                            r['avg_daily'], r['predicted_sales'],
                            r['remaining_stock'], r['consumption_rate'], judge])

    surplus = [r for r in results if r['remaining_stock'] > 0]
    sellout = [r for r in results if r['remaining_stock'] <= 0]

    print(f"\n=== 在庫消化予測サマリ ===")
    print(f"対象品番数: {len(results)}")
    print(f"在庫余剰: {len(surplus)}品番 (合計残在庫: {sum(r['remaining_stock'] for r in surplus):,})")
    print(f"完売見込: {len(sellout)}品番")

    print(f"\n=== 余剰品番一覧（残在庫順） ===")
    print(f"{'品番':<15} {'カテゴリ':<18} {'現在庫':>7} {'日販':>7} {'予測販売':>8} {'残在庫':>8} {'消化率':>7}")
    print("-" * 90)
    for r in surplus:
        print(f"{r['item_code']:<15} {str(r['cat_child']):<18} {r['stock']:>7,} {r['avg_daily']:>7.1f} {r['predicted_sales']:>8,} {r['remaining_stock']:>8,} {r['consumption_rate']:>6.1f}%")

    print(f"\n結果を {output_path} に保存しました。")


def main():
    parser = argparse.ArgumentParser(description='在庫消化予測')
    parser.add_argument('inventory_xlsx', help='品番別在庫Excelファイル')
    parser.add_argument('coefficients_json', help='52週係数JSONファイル')
    parser.add_argument('target_date', help='予測期限 (YYYY-MM-DD)')
    parser.add_argument('--sheet', default=None, help='対象シート名')
    parser.add_argument('--targets', default=None, help='対象品番（カンマ区切り）')
    parser.add_argument('--recent-days', type=int, default=14, help='日販計算の直近日数')
    parser.add_argument('--output', default='forecast_results.csv', help='出力CSVパス')
    args = parser.parse_args()

    target_date = datetime.strptime(args.target_date, '%Y-%m-%d')
    target_set = set(args.targets.split(',')) if args.targets else None

    with open(args.coefficients_json, 'r', encoding='utf-8') as f:
        coefficients = json.load(f)

    wb = openpyxl.load_workbook(args.inventory_xlsx, read_only=True, data_only=True)
    sheet_name = args.sheet or wb.sheetnames[0]
    ws = wb[sheet_name]

    items = load_items(ws, target_set, args.recent_days)
    wb.close()

    if not items:
        print("Error: 対象品番のデータが見つかりません。")
        sys.exit(1)

    # current_dateは最終データ日の翌日
    current_date = items[0]['last_data_date'] + timedelta(days=1)
    print(f"予測開始日: {current_date.strftime('%Y-%m-%d')} (第{get_week_number(current_date)}週)")
    print(f"予測期限: {target_date.strftime('%Y-%m-%d')} (第{get_week_number(target_date)}週)")
    print(f"対象品番数: {len(items)}")

    results = forecast(items, coefficients, current_date, target_date)
    output_results(results, args.output)


if __name__ == '__main__':
    main()
