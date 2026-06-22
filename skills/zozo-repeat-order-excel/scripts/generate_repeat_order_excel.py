#!/usr/bin/env python3
"""
リピート発注管理表Excel自動生成スクリプト（汎用版）

使い方:
  python3 generate_repeat_order_excel.py --item-code sc841 --total-order 6000 --color-mode existing

引数:
  --item-code    : ブランド品番（必須）
  --total-order  : リピート発注総数（必須）
  --color-mode   : existing（既存色のみ）/ all（全色）/ specified（指定色のみ）
  --colors       : 指定色リスト（カンマ区切り、--color-mode specified の場合）
  --output       : 出力ファイルパス（省略時: 発注管理表_{品番}_リピート.xlsx）
  --work-dir     : 作業ディレクトリ（省略時: /home/ubuntu/work）

前提:
  - 作業ディレクトリに以下のファイルが必要:
    - format_10day.xlsm（テンプレートExcel）
    - inventory_{shop}.csv（在庫分析CSV、cp932）
    - color_master.json（ZOZOカラーマスタ）
    - goods_cs_extract_{item_code}.csv（展開SKU CSVからの抽出結果）
    - arrival_data.json（入荷予定データ）
    - reserve_list_{shop}.csv（予約管理一覧CSV、cp932）
  - 注文生データは /home/ubuntu/order_data/{shop}/ に月次CSVとして配置
"""

import argparse
import csv
import io
import json
import os
import subprocess
import sys
from copy import copy
from datetime import datetime

import numpy as np
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import pandas as pd
import requests


# ============================================================
# 引数パース
# ============================================================
def parse_args():
    parser = argparse.ArgumentParser(description='リピート発注管理表Excel自動生成')
    parser.add_argument('--item-code', required=True, help='ブランド品番')
    parser.add_argument('--total-order', required=True, type=int, help='リピート発注総数')
    parser.add_argument('--color-mode', default='existing', choices=['existing', 'all', 'specified'])
    parser.add_argument('--colors', default='', help='指定色リスト（カンマ区切り）')
    parser.add_argument('--output', default='', help='出力ファイルパス')
    parser.add_argument('--work-dir', default='/home/ubuntu/work', help='作業ディレクトリ')
    return parser.parse_args()


# ============================================================
# ショップ名の正規化（フォルダ名用）
# ============================================================
SHOP_FOLDER_MAP = {
    'MONO-MART': 'mono-mart',
    'EMMA CLOTHES': 'emma-clothes',
    'ADRER': 'adrer',
    'CLEL': 'clel',
    'Chaco closet': 'chaco-closet',
    'Anchor Smith': 'anchor-smith',
    'BONLECILL': 'bonlecill',
}


# ============================================================
# 1. 展開SKU CSVから品番情報を取得
# ============================================================
def get_sku_info(item_code, work_dir):
    """展開SKU CSVから品番のSKU情報とショップ名を取得"""
    goods_csv = os.path.join(work_dir, 'goods_cs.csv')
    extract_csv = os.path.join(work_dir, f'goods_cs_extract_{item_code}.csv')

    if not os.path.exists(extract_csv):
        # grepで品番を抽出
        print(f"展開SKU CSVから {item_code} を抽出中...")
        result = subprocess.run(
            ['grep', '-i', item_code, goods_csv],
            capture_output=True, text=True, encoding='cp932', errors='replace'
        )
        header = subprocess.run(
            ['head', '-1', goods_csv],
            capture_output=True, text=True, encoding='cp932', errors='replace'
        ).stdout.strip()

        with open(extract_csv, 'w', encoding='utf-8') as f:
            f.write(header + '\n')
            f.write(result.stdout)

    df = pd.read_csv(extract_csv, encoding='utf-8', dtype=str)
    df.columns = df.columns.str.strip()
    df = df[df['ブランド品番'].str.strip() == item_code]

    if len(df) == 0:
        print(f"エラー: 品番 {item_code} が展開SKU CSVに見つかりません")
        sys.exit(1)

    shop = df['ショップ名'].iloc[0].strip()
    product_code = df['商品コード'].iloc[0].strip()
    item_name = df['商品名'].iloc[0].strip() if '商品名' in df.columns else ''

    print(f"ショップ: {shop}, 商品コード: {product_code}, 商品名: {item_name}")
    return df, shop, product_code, item_name


# ============================================================
# 2. 在庫分析CSVから在庫情報を取得
# ============================================================
def get_inventory(item_code, shop, work_dir):
    """在庫分析CSVから品番の在庫情報を取得"""
    inv_csv = os.path.join(work_dir, f'inventory_{SHOP_FOLDER_MAP.get(shop, shop.lower())}.csv')
    if not os.path.exists(inv_csv):
        # ショップ名そのままでも試す
        inv_csv = os.path.join(work_dir, f'在庫状況({shop}).csv')

    if not os.path.exists(inv_csv):
        print(f"警告: 在庫分析CSV が見つかりません: {inv_csv}")
        return pd.DataFrame(), []

    df = pd.read_csv(inv_csv, encoding='cp932', dtype=str)
    df.columns = df.columns.str.strip()
    df_item = df[df['ブランド品番'].str.strip() == item_code].copy()
    print(f"在庫分析: {len(df_item)} SKU")

    # 在庫情報をレコードリストに変換
    records = []
    for _, r in df_item.iterrows():
        records.append({
            'カラー': r.get('カラー', '').strip(),
            'サイズ': r.get('サイズ', '').strip(),
            'CS品番': r.get('CS品番', '').strip(),
            '在庫': int(r.get('販売可能数', '0').strip() or 0),
            '30日販売': int(r.get('直近30日販売数', '0').strip() or 0),
            '7日販売': int(r.get('直近7日販売数', '0').strip() or 0),
            'お気に入り': int(r.get('お気に入り登録数', '0').strip() or 0),
            '販売価格': int(r.get('販売価格（税抜）', '0').strip() or 0),
            'プロパー価格': int(r.get('プロパー価格（税抜）', '0').strip() or 0),
        })

    return df_item, records


# ============================================================
# 3. SKUバランス配分
# ============================================================
def calc_sku_balance(records, total_order, color_mode='existing', specified_colors=None):
    """SKUバランスを算出して発注数を配分"""
    df = pd.DataFrame(records)

    if color_mode == 'specified' and specified_colors:
        df = df[df['カラー'].isin(specified_colors)].reset_index(drop=True)

    if len(df) == 0:
        print("エラー: 対象SKUが0件です")
        sys.exit(1)

    # サイズ順序
    size_order = {'FREE': 0, 'XS': 1, 'S': 2, 'M': 3, 'L': 4, 'XL': 5, 'XXL': 6, '2XL': 6, '3XL': 7}
    df['サイズ順'] = df['サイズ'].map(size_order).fillna(99)
    df = df.sort_values(['カラー', 'サイズ順']).reset_index(drop=True)

    total_30d = df['30日販売'].sum()
    if total_30d == 0:
        # 30日販売が全て0の場合は均等配分
        df['販売構成比'] = 1 / len(df)
    else:
        df['販売構成比'] = df['30日販売'] / total_30d

    df['初期配分'] = (df['販売構成比'] * total_order).round(0).astype(int)

    # 在庫日数による調整
    df['日販'] = df['30日販売'] / 30
    df['在庫日数'] = np.where(df['日販'] > 0, df['在庫'] / df['日販'], 999)

    df['在庫調整係数'] = np.where(
        df['在庫日数'] <= 10, 1.3,
        np.where(df['在庫日数'] <= 20, 1.15,
        np.where(df['在庫日数'] <= 30, 1.0,
        np.where(df['在庫日数'] <= 60, 0.85,
        0.7)))
    )
    df.loc[df['在庫'] == 0, '在庫調整係数'] = 1.5

    df['調整後配分'] = (df['初期配分'] * df['在庫調整係数']).round(0).astype(int)

    # 合計を total_order に調整
    current_total = df['調整後配分'].sum()
    diff = total_order - current_total
    if diff != 0:
        df_sorted = df.sort_values('30日販売', ascending=False)
        idx_list = df_sorted.index.tolist()
        i = 0
        step = 1 if diff > 0 else -1
        while diff != 0:
            df.loc[idx_list[i % len(idx_list)], '調整後配分'] += step
            diff -= step
            i += 1

    df['最終発注数'] = df['調整後配分']
    print(f"SKUバランス算出完了: {len(df)} SKU, 合計 {df['最終発注数'].sum()} 着")
    return df


# ============================================================
# 4. 注文生データから売上実績を抽出
# ============================================================
def get_order_data(item_code, shop, work_dir):
    """注文生データから品番の注文データを抽出"""
    shop_folder = SHOP_FOLDER_MAP.get(shop, shop.lower())
    order_dir = os.path.join('/home/ubuntu/order_data', shop_folder)

    all_orders = []
    if not os.path.exists(order_dir):
        print(f"警告: 注文データディレクトリが見つかりません: {order_dir}")
        return all_orders

    for fname in sorted(os.listdir(order_dir)):
        if not fname.endswith('.csv'):
            continue
        fpath = os.path.join(order_dir, fname)
        try:
            with open(fpath, 'r', encoding='utf-8') as f:
                header_line = f.readline().strip()

            result = subprocess.run(['grep', item_code, fpath], capture_output=True, text=True)
            if result.stdout.strip():
                for line in result.stdout.strip().split('\n'):
                    reader = csv.reader(io.StringIO(header_line + '\n' + line))
                    rows = list(reader)
                    if len(rows) == 2:
                        header = rows[0]
                        data = rows[1]
                        if len(data) > len(header):
                            data = data[:len(header)]
                        elif len(data) < len(header):
                            data = data + [''] * (len(header) - len(data))
                        row_dict = dict(zip(header, data))
                        if row_dict.get('ブランド品番', '').strip().strip('"') == item_code:
                            all_orders.append(row_dict)
        except Exception:
            pass

    print(f"注文データ: {len(all_orders)} 件")
    return all_orders


# ============================================================
# 5. 入荷予定データを取得
# ============================================================
def get_arrival_data(item_code, work_dir):
    """入荷予定データから品番のデータを取得"""
    arrival_json = os.path.join(work_dir, 'arrival_data.json')
    if not os.path.exists(arrival_json):
        print("警告: 入荷予定データが見つかりません")
        return pd.DataFrame()

    with open(arrival_json, 'r') as f:
        data = json.load(f)

    if isinstance(data, list):
        df = pd.DataFrame(data)
    elif isinstance(data, dict) and 'values' in data:
        values = data['values']
        if len(values) > 1:
            df = pd.DataFrame(values[1:], columns=values[0])
        else:
            return pd.DataFrame()
    else:
        return pd.DataFrame()

    # 品番でフィルタ
    for col in ['ZOZO親品番', 'ブランド品番', '品番']:
        if col in df.columns:
            df_item = df[df[col].astype(str).str.strip() == item_code]
            if len(df_item) > 0:
                print(f"入荷予定: {len(df_item)} 件")
                return df_item

    print("入荷予定: 0 件")
    return pd.DataFrame()


# ============================================================
# 6. 予約管理一覧データを取得
# ============================================================
def get_reserve_data(item_code, shop, work_dir):
    """予約管理一覧CSVから品番のデータを取得"""
    shop_folder = SHOP_FOLDER_MAP.get(shop, shop.lower())
    reserve_csv = os.path.join(work_dir, f'reserve_list_{shop_folder}.csv')
    if not os.path.exists(reserve_csv):
        reserve_csv = os.path.join(work_dir, f'reserve_list_{shop}.csv')

    if not os.path.exists(reserve_csv):
        print("警告: 予約管理一覧CSVが見つかりません")
        return pd.DataFrame()

    df = pd.read_csv(reserve_csv, encoding='cp932', dtype=str)
    df.columns = df.columns.str.strip()
    df_item = df[df['ブランド品番'].str.strip() == item_code]
    print(f"予約管理一覧: {len(df_item)} 件")
    return df_item


# ============================================================
# 7. 商品画像を取得
# ============================================================
def download_images(product_code, colors, work_dir):
    """ZOZOカラーマスタから画像URLを生成しダウンロード"""
    color_master_path = os.path.join(work_dir, 'color_master.json')
    if not os.path.exists(color_master_path):
        print("警告: color_master.json が見つかりません。画像はスキップします。")
        return {}

    with open(color_master_path, 'r') as f:
        color_master = json.load(f)

    # カラー名→カラーコードのマッピング
    color_code_map = {}
    if isinstance(color_master, list):
        for entry in color_master:
            name = entry.get('カラー名', entry.get('name', '')).strip()
            code = entry.get('カラーコード', entry.get('code', '')).strip()
            if name and code:
                color_code_map[name] = code
    elif isinstance(color_master, dict):
        color_code_map = color_master

    img_dir = os.path.join(work_dir, 'images')
    os.makedirs(img_dir, exist_ok=True)

    suffix = product_code[-3:]
    image_paths = {}

    for color in colors:
        code = color_code_map.get(color, '')
        if not code:
            print(f"  カラーコード不明: {color}")
            continue

        url = f"https://o.imgz.jp/{suffix}/{product_code}/{product_code}b_{code}_d.jpg"
        img_path = os.path.join(img_dir, f'{product_code}_{code}.jpg')

        if os.path.exists(img_path):
            image_paths[color] = img_path
            continue

        try:
            resp = requests.get(url, timeout=10)
            if resp.status_code == 200:
                with open(img_path, 'wb') as f:
                    f.write(resp.content)
                image_paths[color] = img_path
                print(f"  画像取得: {color} ({code})")
            else:
                print(f"  画像取得失敗: {color} ({code}) -> {resp.status_code}")
        except Exception as e:
            print(f"  画像取得エラー: {color} -> {e}")

    return image_paths


# ============================================================
# 8. 10日刻み集計
# ============================================================
def calc_10day_sales(all_orders, item_code):
    """注文データから10日刻みの売上集計を作成"""
    if not all_orders:
        return pd.DataFrame(), pd.DataFrame()

    order_df = pd.DataFrame(all_orders)
    order_df['ブランド品番'] = order_df['ブランド品番'].str.strip().str.strip('"')
    order_df['CS品番'] = order_df['CS品番'].str.strip().str.strip('"')
    order_df['注文数'] = pd.to_numeric(order_df['注文数'].str.strip().str.strip('"'), errors='coerce').fillna(0).astype(int)
    order_df['注文日'] = order_df['注文日'].str.strip().str.strip('"')
    order_df['販売タイプ'] = order_df['販売タイプ'].str.strip().str.strip('"')
    order_df['SKU'] = order_df['ブランド品番'] + order_df['CS品番']

    def parse_date(d):
        try:
            d = str(d).strip().strip('"')
            if ' ' in d:
                d = d.split(' ')[0]
            if '/' in d:
                parts = d.split('/')
            elif '-' in d:
                parts = d.split('-')
            else:
                return None, None, None
            return int(parts[0]), int(parts[1]), int(parts[2])
        except:
            return None, None, None

    order_df[['年', '月', '日']] = order_df['注文日'].apply(lambda x: pd.Series(parse_date(x)))
    order_df = order_df.dropna(subset=['年', '月', '日'])
    order_df['年'] = order_df['年'].astype(int)
    order_df['月'] = order_df['月'].astype(int)
    order_df['日'] = order_df['日'].astype(int)

    def get_10day(day):
        if day <= 10:
            return '1~10'
        elif day <= 20:
            return '11~20'
        else:
            return '21~31'

    order_df['10日刻み'] = order_df['日'].apply(get_10day)

    pivot = order_df.groupby(['SKU', '年', '月', '10日刻み'])['注文数'].sum().reset_index()
    return order_df, pivot


# ============================================================
# 9. 発注表の列番号算出
# ============================================================
def get_col_for_period(year, month, ten_day):
    """年月10日刻みから発注表の列番号を返す"""
    ten_day_labels = ['1~10', '11~20', '21~31']
    if year == 2023:
        month_offset = month - 1
    elif year == 2024:
        month_offset = 12 + (month - 1)
    elif year == 2025:
        month_offset = 24 + (month - 1)
    elif year == 2026:
        month_offset = 36 + (month - 1)
    else:
        return None

    ten_day_offset = ten_day_labels.index(ten_day) if ten_day in ten_day_labels else 0
    col = 33 + month_offset * 3 + ten_day_offset  # AG=33
    return col if col <= 158 else None  # FB=158まで


# ============================================================
# メイン処理
# ============================================================
def main():
    args = parse_args()
    item_code = args.item_code
    total_order = args.total_order
    color_mode = args.color_mode
    specified_colors = [c.strip() for c in args.colors.split(',') if c.strip()] if args.colors else []
    work_dir = args.work_dir
    output_file = args.output or os.path.join(work_dir, f'発注管理表_{item_code}_リピート.xlsx')

    print(f"=== リピート発注管理表生成 ===")
    print(f"品番: {item_code}, 発注総数: {total_order}, カラーモード: {color_mode}")

    # --- データ収集 ---
    print("\n--- Step 1: SKU情報取得 ---")
    sku_df, shop, product_code, item_name = get_sku_info(item_code, work_dir)

    print("\n--- Step 2: 在庫分析取得 ---")
    inv_full_df, inv_records = get_inventory(item_code, shop, work_dir)

    print("\n--- Step 3: SKUバランス算出 ---")
    balance_df = calc_sku_balance(inv_records, total_order, color_mode, specified_colors)

    print("\n--- Step 4: 注文データ取得 ---")
    all_orders = get_order_data(item_code, shop, work_dir)

    print("\n--- Step 5: 入荷予定取得 ---")
    arrival_df = get_arrival_data(item_code, work_dir)

    print("\n--- Step 6: 予約管理一覧取得 ---")
    reserve_df = get_reserve_data(item_code, shop, work_dir)

    print("\n--- Step 7: 商品画像取得 ---")
    colors = balance_df['カラー'].unique().tolist()
    image_paths = download_images(product_code, colors, work_dir)

    # --- Excel生成 ---
    print("\n--- Step 8: Excel生成 ---")
    template_path = os.path.join(work_dir, 'format_10day.xlsm')
    if not os.path.exists(template_path):
        print(f"エラー: テンプレートExcelが見つかりません: {template_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(template_path, keep_vba=False)

    # --- 発注表シート ---
    ws = wb['発注表']
    ws.cell(row=12, column=5, value=item_code)

    start_row = 14
    for idx, row in balance_df.iterrows():
        r = start_row + idx
        ws.cell(row=r, column=1, value=idx + 1)          # A: 連番
        ws.cell(row=r, column=2, value=item_code)         # B: メーカー品番
        ws.cell(row=r, column=3, value=row['カラー'])     # C: メーカーカラー名
        ws.cell(row=r, column=5, value=item_code)         # E: 品番
        ws.cell(row=r, column=6, value=row['カラー'])     # F: カラー
        ws.cell(row=r, column=7, value=row['サイズ'])     # G: サイズ
        ws.cell(row=r, column=8, value=row['CS品番'])     # H: CS品番
        ws.cell(row=r, column=12, value=int(row['在庫'])) # L: 在庫
        ws.cell(row=r, column=15, value=int(row['30日販売']))  # O: 販売数（30日）
        ws.cell(row=r, column=17, value=int(row['お気に入り']))  # Q: お気に入り
        ws.cell(row=r, column=18, value=f"{item_code}{row['CS品番']}")  # R: SKU
        ws.cell(row=r, column=19, value=int(row['7日販売']))  # S: 7日間
        ws.cell(row=r, column=20, value=int(row['30日販売']))  # T: 一ヶ月
        ws.cell(row=r, column=21, value=int(row['最終発注数']))  # U: 発注数
        ws.cell(row=r, column=22, value=int(row['在庫']))  # V: フリー在庫
        daily_7d = row['7日販売'] / 7 if row['7日販売'] > 0 else 0
        ws.cell(row=r, column=23, value=round(row['在庫'] / daily_7d, 1) if daily_7d > 0 else 0)  # W
        daily_30d = row['30日販売'] / 30 if row['30日販売'] > 0 else 0
        ws.cell(row=r, column=24, value=round(row['在庫'] / daily_30d, 1) if daily_30d > 0 else 0)  # X
        ws.cell(row=r, column=27, value=0)  # AA: 予約未処理数
        ws.cell(row=r, column=28, value=int(row['お気に入り']))  # AB: お気に入り

    last_data_row = start_row + len(balance_df) - 1
    print(f"発注表: {len(balance_df)} SKU 書き込み完了")

    # --- 10日刻み売上実績 ---
    print("10日刻み売上実績を書き込み中...")
    order_df, pivot = calc_10day_sales(all_orders, item_code)
    if len(pivot) > 0:
        for idx, row in balance_df.iterrows():
            r = start_row + idx
            sku = f"{item_code}{row['CS品番']}"
            sku_data = pivot[pivot['SKU'] == sku]
            for _, period in sku_data.iterrows():
                col = get_col_for_period(int(period['年']), int(period['月']), period['10日刻み'])
                if col:
                    ws.cell(row=r, column=col, value=int(period['注文数']))

        # Row7（予約）とRow8（実売）
        reserve_orders = order_df[order_df['販売タイプ'] == '予約']
        normal_orders = order_df[order_df['販売タイプ'] != '予約']
        for orders_subset, target_row in [(reserve_orders, 7), (normal_orders, 8)]:
            if len(orders_subset) > 0:
                period_sum = orders_subset.groupby(['年', '月', '10日刻み'])['注文数'].sum().reset_index()
                for _, period in period_sum.iterrows():
                    col = get_col_for_period(int(period['年']), int(period['月']), period['10日刻み'])
                    if col:
                        ws.cell(row=target_row, column=col, value=int(period['注文数']))
        print(f"10日刻み: {len(pivot)} レコード書き込み完了")

    # --- 年度別平均売価 ---
    if len(order_df) > 0:
        order_df['販売価格（税抜）'] = pd.to_numeric(
            order_df['販売価格（税抜）'].str.strip().str.strip('"'), errors='coerce')
        order_df['合計金額（税抜）'] = pd.to_numeric(
            order_df['合計金額（税抜）'].str.strip().str.strip('"'), errors='coerce')
        for idx, row in balance_df.iterrows():
            r = start_row + idx
            sku = f"{item_code}{row['CS品番']}"
            sku_orders = order_df[order_df['SKU'] == sku]
            for year, col in [(2023, 29), (2024, 30), (2025, 31), (2026, 32)]:
                year_orders = sku_orders[sku_orders['年'] == year]
                if len(year_orders) > 0 and year_orders['注文数'].sum() > 0:
                    total_amount = year_orders['合計金額（税抜）'].sum()
                    total_qty = year_orders['注文数'].sum()
                    avg_price = round(total_amount / total_qty) if total_qty > 0 else 0
                    ws.cell(row=r, column=col, value=avg_price)

    # --- 売上実績シート ---
    print("売上実績シートを書き込み中...")
    ws_sales = wb['売上実績']
    sales_cols = [
        'ショップ名', '親カテゴリ', '子カテゴリ', '親商品タイプ', '子商品タイプ', '性別',
        'ブランド品番', '商品名', 'CS品番', 'カラー', 'サイズ', '販売価格（税抜）',
        '販売タイプ', '価格タイプ', 'プロパー価格（税抜）', '注文数', '合計金額（税抜）',
        '注文日', 'バーコード'
    ]
    for col_idx, col_name in enumerate(sales_cols, 1):
        ws_sales.cell(row=1, column=col_idx, value=col_name)
    calc_headers = ['SKU', '年', '月', '日', '前後半', '10日刻み', '5日ごと', '3日ごと', '日付変換']
    for col_idx, col_name in enumerate(calc_headers, 20):
        ws_sales.cell(row=1, column=col_idx, value=col_name)

    row_num = 2
    for order in all_orders:
        for col_idx, col_name in enumerate(sales_cols, 1):
            val = order.get(col_name, '').strip().strip('"')
            if col_name in ['販売価格（税抜）', 'プロパー価格（税抜）', '注文数', '合計金額（税抜）']:
                try:
                    val = int(val)
                except:
                    pass
            ws_sales.cell(row=row_num, column=col_idx, value=val)

        brand_code = order.get('ブランド品番', '').strip().strip('"')
        cs_code = order.get('CS品番', '').strip().strip('"')
        ws_sales.cell(row=row_num, column=20, value=f"{brand_code}{cs_code}")

        order_date = order.get('注文日', '').strip().strip('"')
        try:
            if ' ' in order_date:
                order_date = order_date.split(' ')[0]
            if '/' in order_date:
                parts = order_date.split('/')
            elif '-' in order_date:
                parts = order_date.split('-')
            else:
                parts = []
            if len(parts) >= 3:
                year, month, day = int(parts[0]), int(parts[1]), int(parts[2])
                ws_sales.cell(row=row_num, column=21, value=str(year))
                ws_sales.cell(row=row_num, column=22, value=month)
                ws_sales.cell(row=row_num, column=23, value=day)
                ws_sales.cell(row=row_num, column=24, value='前半' if day <= 15 else '後半')
                if day <= 10:
                    ten_day = '1~10'
                elif day <= 20:
                    ten_day = '11~20'
                else:
                    ten_day = '21~31'
                ws_sales.cell(row=row_num, column=25, value=ten_day)
                ws_sales.cell(row=row_num, column=26, value=f"{((day-1)//5)*5+1}~{min(((day-1)//5)*5+5, 31)}")
                ws_sales.cell(row=row_num, column=27, value=f"{((day-1)//3)*3+1}~{min(((day-1)//3)*3+3, 31)}")
                try:
                    ws_sales.cell(row=row_num, column=28, value=datetime(year, month, day))
                except:
                    pass
        except:
            pass
        row_num += 1
    print(f"売上実績: {row_num - 2} 行書き込み完了")

    # --- 在庫分析貼り付けシート ---
    print("在庫分析シートを書き込み中...")
    ws_inv = wb['在庫分析貼り付け']
    if len(inv_full_df) > 0:
        for col_idx, col_name in enumerate(inv_full_df.columns, 1):
            ws_inv.cell(row=1, column=col_idx, value=col_name)
        for row_idx, (_, row) in enumerate(inv_full_df.iterrows(), 2):
            for col_idx, val in enumerate(row, 1):
                ws_inv.cell(row=row_idx, column=col_idx, value=val)
        print(f"在庫分析: {len(inv_full_df)} 行書き込み完了")

    # --- 予約管理表シート ---
    print("予約管理表シートを書き込み中...")
    ws_reserve = wb['予約管理表']
    if len(arrival_df) > 0:
        for col_idx, col_name in enumerate(arrival_df.columns, 1):
            ws_reserve.cell(row=1, column=col_idx, value=col_name)
        for row_idx, (_, row) in enumerate(arrival_df.iterrows(), 2):
            for col_idx, val in enumerate(row, 1):
                if pd.notna(val):
                    ws_reserve.cell(row=row_idx, column=col_idx, value=val)
        print(f"予約管理表: {len(arrival_df)} 行書き込み完了")

    # --- 予約一覧シート ---
    print("予約一覧シートを書き込み中...")
    ws_yoyaku = wb['予約一覧']
    reserve_headers = [
        '表示', 'ショップ', '親カテゴリ', '商品名', '商品コード', 'ブランド品番',
        'カラー', 'サイズ', 'CS品番', '販売価格（税抜）', '販売可能数', '在庫数',
        '予約受付数', '注文数', '未処理', '発送指定日', 'お届け予定(初回納期)',
        '遅延メール送信履歴', 'お届け予定(遅延納期)', '販売終了', 'バーコード',
        '自動配信禁止設定', '納期メモ'
    ]
    for col_idx, col_name in enumerate(reserve_headers, 1):
        ws_yoyaku.cell(row=1, column=col_idx, value=col_name)
    if len(reserve_df) > 0:
        for row_idx, (_, row) in enumerate(reserve_df.iterrows(), 2):
            for col_idx, col_name in enumerate(reserve_headers, 1):
                val = row.get(col_name, '')
                if pd.notna(val):
                    ws_yoyaku.cell(row=row_idx, column=col_idx, value=val)
        print(f"予約一覧: {len(reserve_df)} 行書き込み完了")
    else:
        print("予約一覧: 0 件")

    # --- 商品画像埋め込み ---
    print("商品画像を埋め込み中...")
    if image_paths:
        img_start_row = last_data_row + 3
        for i, color in enumerate(colors):
            if color in image_paths:
                try:
                    img = Image(image_paths[color])
                    img.width = 120
                    img.height = 150
                    col = 2 + i * 2
                    cell_ref = f'{get_column_letter(col)}{img_start_row}'
                    ws.add_image(img, cell_ref)
                    ws.cell(row=img_start_row - 1, column=col, value=color)
                    print(f"  画像追加: {color} -> {cell_ref}")
                except Exception as e:
                    print(f"  画像追加エラー: {color} -> {e}")

    # --- 保存 ---
    print(f"\n--- 保存 ---")
    wb.save(output_file)
    wb.close()
    file_size = os.path.getsize(output_file) / 1024 / 1024
    print(f"保存完了: {output_file} ({file_size:.1f} MB)")


if __name__ == '__main__':
    main()
